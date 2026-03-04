from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QTableWidget, QTableWidgetItem,
                             QComboBox, QGroupBox, QMessageBox,
                             QHeaderView, QTabWidget, QTextEdit,
                             QSplitter, QFrame, QGridLayout,
                             QCheckBox, QScrollArea, QDateEdit,
                             QFileDialog, QProgressDialog, QProgressBar,
                             QLineEdit)  # ADICIONADO QLineEdit
from PyQt6.QtCore import Qt, QTimer, QDate, QCoreApplication, QProcess
from PyQt6.QtGui import QFont, QColor
import sqlite3
from datetime import datetime, timedelta
import json
import pandas as pd
import os
import subprocess
import sys
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

class TelaRelatorios(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Título
        titulo = QLabel("📊 RELATÓRIOS E ESTATÍSTICAS - USINA BETTO MIX")
        titulo.setFont(QFont("Arial", 18, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("""
            color: white;
            background-color: #0D47A1;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 10px;
        """)
        layout.addWidget(titulo)
        
        # Abas
        self.tabs = QTabWidget()
        
        # Aba 1: Consumo Diário (COM FILTROS DIA, MÊS, ANO)
        tab_diario = self.criar_aba_consumo_diario()
        self.tabs.addTab(tab_diario, "📅 DIÁRIO")
        
        # Aba 2: Produção (COM MOTORISTA, PLACA E OBSERVAÇÃO ADICIONADOS) - COM FILTROS MELHORADOS
        tab_producao = self.criar_aba_producao_com_filtros()
        self.tabs.addTab(tab_producao, "⚙️ PRODUÇÃO")
        
        # Aba 3: Semanal COM FILTROS DIA, MÊS, ANO
        tab_semanal = self.criar_aba_semanal()
        self.tabs.addTab(tab_semanal, "📈 SEMANAL")
        
        # Aba 4: Mensal
        tab_mensal = self.criar_aba_mensal()
        self.tabs.addTab(tab_mensal, "📊 MENSAL")
        
        # Aba 5: Gastos (CORRIGIDA E SINCRONIZADA)
        tab_gastos = self.criar_aba_gastos()
        self.tabs.addTab(tab_gastos, "💰 GASTOS")
        
        # Aba 6: Estoque Detalhado
        tab_estoque = self.criar_aba_estoque_detalhado()
        self.tabs.addTab(tab_estoque, "📦 ESTOQUE")
        
        layout.addWidget(self.tabs)
        
        # Botões de ação
        btn_layout = QHBoxLayout()
        
        btn_gerar = QPushButton("🔄 Atualizar Todos")
        btn_gerar.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_gerar.clicked.connect(self.gerar_relatorios)
        
        btn_exportar = QPushButton("📤 Exportar Excel")
        btn_exportar.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_exportar.clicked.connect(self.exportar_excel)
        
        btn_imprimir = QPushButton("🖨️ Imprimir Relatório")
        btn_imprimir.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_imprimir.clicked.connect(self.imprimir_relatorio_corrigido)
        
        btn_layout.addWidget(btn_gerar)
        btn_layout.addWidget(btn_exportar)
        btn_layout.addWidget(btn_imprimir)
        btn_layout.addStretch()
        
        layout.addLayout(btn_layout)
    
    def criar_aba_consumo_diario(self):
        """Cria aba de consumo diário com filtros por dia, mês e ano"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Frame superior - Controles e seleção
        frame_controles = QFrame()
        frame_controles.setFrameStyle(QFrame.Shape.StyledPanel)
        frame_controles.setMaximumHeight(100)
        controles_layout = QHBoxLayout(frame_controles)
        
        # Frame para filtros
        frame_filtros = QFrame()
        frame_filtros.setFrameStyle(QFrame.Shape.Box)
        frame_filtros.setStyleSheet("""
            QFrame {
                border: 1px solid #1976D2;
                border-radius: 5px;
                padding: 5px;
                background-color: #E3F2FD;
            }
        """)
        
        filtros_layout = QGridLayout(frame_filtros)
        
        # Filtro por Dia
        filtros_layout.addWidget(QLabel("Dia:"), 0, 0)
        self.combo_dia_filtro = QComboBox()
        self.combo_dia_filtro.addItem("Todos os dias")
        for dia in range(1, 32):
            self.combo_dia_filtro.addItem(f"{dia:02d}")
        self.combo_dia_filtro.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_dia_filtro, 0, 1)
        
        # Filtro por Mês
        filtros_layout.addWidget(QLabel("Mês:"), 0, 2)
        self.combo_mes_filtro = QComboBox()
        self.combo_mes_filtro.addItem("Todos os meses")
        meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        for i, mes in enumerate(meses, 1):
            self.combo_mes_filtro.addItem(mes, i)
        self.combo_mes_filtro.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_mes_filtro, 0, 3)
        
        # Filtro por Ano
        filtros_layout.addWidget(QLabel("Ano:"), 0, 4)
        self.combo_ano_filtro = QComboBox()
        self.combo_ano_filtro.addItem("Todos os anos")
        ano_atual = datetime.now().year
        for ano in range(2020, 2826):
            self.combo_ano_filtro.addItem(str(ano))
        self.combo_ano_filtro.setCurrentText(str(ano_atual))
        self.combo_ano_filtro.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_ano_filtro, 0, 5)
        
        controles_layout.addWidget(frame_filtros)
        
        # Botões
        btn_container = QFrame()
        btn_container_layout = QVBoxLayout(btn_container)
        
        btn_gerar_diario = QPushButton("📊 Gerar Relatório")
        btn_gerar_diario.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                min-width: 150px;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_gerar_diario.clicked.connect(self.gerar_relatorio_diario)
        
        btn_limpar_filtros = QPushButton("🗑️ Limpar")
        btn_limpar_filtros.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 5px 10px;
                border-radius: 5px;
                font-weight: bold;
                margin-top: 5px;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_limpar_filtros.clicked.connect(self.limpar_filtros_diario)
        
        btn_container_layout.addWidget(btn_gerar_diario)
        btn_container_layout.addWidget(btn_limpar_filtros)
        
        controles_layout.addWidget(btn_container)
        
        # Checkbox para mostrar todos os materiais
        self.check_todos_materiais = QCheckBox("Mostrar todos os materiais")
        self.check_todos_materiais.setChecked(True)
        self.check_todos_materiais.stateChanged.connect(self.gerar_relatorio_diario)
        controles_layout.addWidget(self.check_todos_materiais)
        
        layout.addWidget(frame_controles)
        
        # Splitter para dividir a tela
        splitter = QSplitter(Qt.Orientation.Vertical)
        
        # Painel superior - Volume e resumo
        painel_superior = QFrame()
        painel_superior.setFrameStyle(QFrame.Shape.StyledPanel)
        layout_superior = QVBoxLayout(painel_superior)
        
        # Informações de produção do dia/período
        self.group_producao = QGroupBox("📦 PRODUÇÃO DO PERÍODO")
        self.group_producao.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #E3F2FD;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        layout_producao = QGridLayout()
        
        # Labels para mostrar os dados
        self.label_volume_total = QLabel("Volume total carregado: 0.00 m³")
        self.label_volume_total.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        
        self.label_num_pesagens = QLabel("Número de pesagens: 0")
        self.label_num_pesagens.setFont(QFont("Arial", 11))
        
        self.label_tracos_utilizados = QLabel("Traços utilizados: 0")
        self.label_tracos_utilizados.setFont(QFont("Arial", 11))
        
        self.label_clientes_atendidos = QLabel("Clientes atendidos: 0")
        self.label_clientes_atendidos.setFont(QFont("Arial", 11))
        
        layout_producao.addWidget(self.label_volume_total, 0, 0, 1, 2)
        layout_producao.addWidget(self.label_num_pesagens, 1, 0)
        layout_producao.addWidget(self.label_tracos_utilizados, 1, 1)
        layout_producao.addWidget(self.label_clientes_atendidos, 2, 0, 1, 2)
        
        self.group_producao.setLayout(layout_producao)
        layout_superior.addWidget(self.group_producao)
        
        # Saldos atuais dos materiais
        self.group_saldos = QGroupBox("💰 SALDOS ATUAIS DOS MATERIAIS")
        self.group_saldos.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #E3F2FD;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        layout_saldos = QGridLayout()
        
        # Lista de todos os materiais para mostrar saldos
        self.materiais_labels = {}
        materiais = [
            ("Cimento", "kg", "#1565C0"),
            ("Brita 0", "kg", "#0D47A1"),
            ("Brita 1", "kg", "#1976D2"),
            ("Areia Média", "kg", "#2196F3"),
            ("Pó de Brita", "kg", "#64B5F6"),
            ("Aditivo", "kg", "#1E88E5"),
            ("Água", "litros", "#42A5F5")
        ]
        
        for i, (material, unidade, cor) in enumerate(materiais):
            frame_material = QFrame()
            frame_material.setFrameStyle(QFrame.Shape.Box)
            frame_material.setStyleSheet(f"""
                QFrame {{
                    border: 2px solid {cor};
                    border-radius: 5px;
                    padding: 5px;
                    background-color: white;
                }}
            """)
            
            layout_material = QVBoxLayout(frame_material)
            
            label_nome = QLabel(f"<b>{material}</b>")
            label_nome.setAlignment(Qt.AlignmentFlag.AlignCenter)
            
            label_valor = QLabel("0.00")
            label_valor.setFont(QFont("Arial", 16, QFont.Weight.Bold))
            label_valor.setAlignment(Qt.AlignmentFlag.AlignCenter)
            label_valor.setStyleSheet(f"color: {cor};")
            
            label_unidade = QLabel(unidade)
            label_unidade.setAlignment(Qt.AlignmentFlag.AlignCenter)
            label_unidade.setStyleSheet("color: #666;")
            
            layout_material.addWidget(label_nome)
            layout_material.addWidget(label_valor)
            layout_material.addWidget(label_unidade)
            
            # Armazenar referência para atualização
            self.materiais_labels[material] = label_valor
            
            layout_saldos.addWidget(frame_material, i // 4, i % 4)
        
        self.group_saldos.setLayout(layout_saldos)
        layout_superior.addWidget(self.group_saldos)
        
        painel_superior.setLayout(layout_superior)
        
        # Painel inferior - Tabela detalhada
        painel_inferior = QFrame()
        painel_inferior.setFrameStyle(QFrame.Shape.StyledPanel)
        layout_inferior = QVBoxLayout(painel_inferior)
        
        # Tabela de consumo
        self.tabela_diario = QTableWidget()
        self.tabela_diario.setColumnCount(6)
        self.tabela_diario.setHorizontalHeaderLabels([
            "Material", "Unidade", "Estoque Inicial", "Consumo", "Estoque Final", "Variação %"
        ])
        
        # Configurar largura das colunas
        header = self.tabela_diario.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        
        # Aplicar estilo da tabela com paleta de azuis
        self.tabela_diario.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #E3F2FD;
                selection-background-color: #BBDEFB;
                border: 1px solid #1976D2;
                gridline-color: #BBDEFB;
            }
            QHeaderView::section {
                background-color: #0D47A1;
                color: white;
                padding: 5px;
                border: 1px solid #0D47A1;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        layout_inferior.addWidget(self.tabela_diario)
        
        # Resumo
        self.label_resumo_diario = QLabel("Selecione os filtros para gerar o relatório")
        self.label_resumo_diario.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label_resumo_diario.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        self.label_resumo_diario.setStyleSheet("color: #0D47A1; padding: 10px; background-color: #E3F2FD; border-radius: 5px; border: 1px solid #1976D2;")
        layout_inferior.addWidget(self.label_resumo_diario)
        
        painel_inferior.setLayout(layout_inferior)
        
        # Adicionar painéis ao splitter
        splitter.addWidget(painel_superior)
        splitter.addWidget(painel_inferior)
        splitter.setSizes([300, 500])
        
        layout.addWidget(splitter)
        
        # Conectar os filtros para atualização automática (SINCRONIZAÇÃO)
        self.combo_dia_filtro.currentIndexChanged.connect(self.gerar_relatorio_diario)
        self.combo_mes_filtro.currentIndexChanged.connect(self.gerar_relatorio_diario)
        self.combo_ano_filtro.currentIndexChanged.connect(self.gerar_relatorio_diario)
        
        # Carregar dados iniciais
        QTimer.singleShot(100, self.carregar_dados_iniciais_diario)
        
        return tab
    
    def carregar_dados_iniciais_diario(self):
        """Carrega dados iniciais da aba diária"""
        hoje = datetime.now()
        self.combo_dia_filtro.setCurrentText(f"{hoje.day:02d}")
        self.combo_mes_filtro.setCurrentIndex(hoje.month)
        self.combo_ano_filtro.setCurrentText(str(hoje.year))
        self.gerar_relatorio_diario()
    
    def limpar_filtros_diario(self):
        """Limpa todos os filtros da aba diária"""
        self.combo_dia_filtro.setCurrentIndex(0)
        self.combo_mes_filtro.setCurrentIndex(0)
        self.combo_ano_filtro.setCurrentIndex(0)
        self.gerar_relatorio_diario()
    
    def criar_aba_producao_com_filtros(self):
        """Cria aba de produção COM SISTEMA DE FILTRO COMPLETO"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Título da aba
        titulo_producao = QLabel("📊 RELATÓRIO DE PRODUÇÃO - CONCRETO CARREGADO")
        titulo_producao.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        titulo_producao.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo_producao.setStyleSheet("""
            color: white;
            background-color: #0D47A1;
            padding: 10px;
            border-radius: 5px;
        """)
        layout.addWidget(titulo_producao)
        
        # Frame de controles - FILTROS COMPLETOS
        frame_filtros = QGroupBox("🔍 FILTROS DE CONSULTA AVANÇADOS")
        frame_filtros.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                margin-top: 5px;
                padding-top: 10px;
                background-color: #E3F2FD;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        filtros_layout = QGridLayout(frame_filtros)
        
        # Linha 1: Filtros básicos
        # Data Inicial
        filtros_layout.addWidget(QLabel("Data Inicial:"), 0, 0)
        self.date_inicial_producao = QDateEdit()
        self.date_inicial_producao.setDisplayFormat("dd/MM/yyyy")
        self.date_inicial_producao.setDate(QDate.currentDate().addDays(-7))
        self.date_inicial_producao.setCalendarPopup(True)
        self.date_inicial_producao.setStyleSheet("""
            QDateEdit {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QDateEdit:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.date_inicial_producao, 0, 1)
        
        # Data Final
        filtros_layout.addWidget(QLabel("Data Final:"), 0, 2)
        self.date_final_producao = QDateEdit()
        self.date_final_producao.setDisplayFormat("dd/MM/yyyy")
        self.date_final_producao.setDate(QDate.currentDate())
        self.date_final_producao.setCalendarPopup(True)
        self.date_final_producao.setStyleSheet("""
            QDateEdit {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QDateEdit:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.date_final_producao, 0, 3)
        
        # Status
        filtros_layout.addWidget(QLabel("Status:"), 0, 4)
        self.combo_status_producao = QComboBox()
        self.combo_status_producao.addItem("Todos")
        self.combo_status_producao.addItem("CONCLUÍDO")
        self.combo_status_producao.addItem("PENDENTE")
        self.combo_status_producao.addItem("CANCELADO")
        self.combo_status_producao.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_status_producao, 0, 5)
        
        # Linha 2: Filtros de motorista e placa
        filtros_layout.addWidget(QLabel("Motorista:"), 1, 0)
        self.combo_motorista_producao = QComboBox()
        self.combo_motorista_producao.addItem("Todos")
        self.combo_motorista_producao.setEditable(True)
        self.combo_motorista_producao.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_motorista_producao, 1, 1)
        
        filtros_layout.addWidget(QLabel("Placa:"), 1, 2)
        self.combo_placa_producao = QComboBox()
        self.combo_placa_producao.addItem("Todas")
        self.combo_placa_producao.setEditable(True)
        self.combo_placa_producao.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_placa_producao, 1, 3)
        
        # Cliente
        filtros_layout.addWidget(QLabel("Cliente:"), 1, 4)
        self.combo_cliente_producao = QComboBox()
        self.combo_cliente_producao.addItem("Todos")
        self.combo_cliente_producao.setEditable(True)
        self.combo_cliente_producao.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_cliente_producao, 1, 5)
        
        # Linha 3: Filtros de traço e observação
        filtros_layout.addWidget(QLabel("Traço:"), 2, 0)
        self.combo_traco_producao = QComboBox()
        self.combo_traco_producao.addItem("Todos")
        self.combo_traco_producao.setEditable(True)
        self.combo_traco_producao.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_traco_producao, 2, 1)
        
        filtros_layout.addWidget(QLabel("Observação:"), 2, 2)
        self.edit_observacao_producao = QLineEdit()
        self.edit_observacao_producao.setPlaceholderText("Buscar na observação...")
        self.edit_observacao_producao.setStyleSheet("""
            QLineEdit {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QLineEdit:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.edit_observacao_producao, 2, 3)
        
        # Volume Mínimo
        filtros_layout.addWidget(QLabel("Volume Mín. (m³):"), 2, 4)
        self.spin_volume_min = QLineEdit()
        self.spin_volume_min.setPlaceholderText("0.00")
        self.spin_volume_min.setStyleSheet("""
            QLineEdit {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QLineEdit:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.spin_volume_min, 2, 5)
        
        # Botões de ação dos filtros
        btn_aplicar = QPushButton("🔍 Aplicar Filtros")
        btn_aplicar.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_aplicar.clicked.connect(self.carregar_dados_producao_com_filtros)
        filtros_layout.addWidget(btn_aplicar, 3, 0, 1, 2)
        
        btn_limpar = QPushButton("🗑️ Limpar Filtros")
        btn_limpar.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_limpar.clicked.connect(self.limpar_filtros_producao_avancados)
        filtros_layout.addWidget(btn_limpar, 3, 2, 1, 2)
        
        btn_carregar_opcoes = QPushButton("🔄 Carregar Opções")
        btn_carregar_opcoes.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_carregar_opcoes.clicked.connect(self.carregar_opcoes_filtros_producao)
        filtros_layout.addWidget(btn_carregar_opcoes, 3, 4, 1, 2)
        
        layout.addWidget(frame_filtros)
        
        # Tabela de produção
        self.tabela_producao_filtrada = QTableWidget()
        self.tabela_producao_filtrada.setColumnCount(11)
        self.tabela_producao_filtrada.setHorizontalHeaderLabels([
            "Data", "Volume (m³)", "Pesagem", "Traço", "Cliente", 
            "M³ Médio", "Status", "Hora", "Motorista", "Placa", "Observação"
        ])
        
        # Configurar largura das colunas
        header = self.tabela_producao_filtrada.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.Stretch)
        
        # Aplicar estilo da tabela
        self.tabela_producao_filtrada.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #E3F2FD;
                selection-background-color: #BBDEFB;
                border: 1px solid #1976D2;
                gridline-color: #BBDEFB;
            }
            QHeaderView::section {
                background-color: #0D47A1;
                color: white;
                padding: 5px;
                border: 1px solid #0D47A1;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        layout.addWidget(self.tabela_producao_filtrada)
        
        # Resumo da produção filtrada
        self.frame_resumo_producao_filtrada = QFrame()
        self.frame_resumo_producao_filtrada.setFrameStyle(QFrame.Shape.StyledPanel)
        self.frame_resumo_producao_filtrada.setStyleSheet("""
            background-color: #E3F2FD;
            border: 2px solid #0D47A1;
            border-radius: 5px;
            padding: 10px;
        """)
        
        resumo_layout = QHBoxLayout(self.frame_resumo_producao_filtrada)
        
        self.label_total_volume_filtrado = QLabel("📦 Volume Total: 0.00 m³")
        self.label_total_volume_filtrado.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        
        self.label_total_pesagens_filtrado = QLabel("⚖️ Total de Pesagens: 0")
        self.label_total_pesagens_filtrado.setFont(QFont("Arial", 11))
        
        self.label_media_m3_filtrado = QLabel("📊 M³ Médio: 0.00")
        self.label_media_m3_filtrado.setFont(QFont("Arial", 11))
        
        self.label_status_geral_filtrado = QLabel("✅ Status: Sem dados")
        self.label_status_geral_filtrado.setFont(QFont("Arial", 11))
        
        resumo_layout.addWidget(self.label_total_volume_filtrado)
        resumo_layout.addWidget(self.label_total_pesagens_filtrado)
        resumo_layout.addWidget(self.label_media_m3_filtrado)
        resumo_layout.addWidget(self.label_status_geral_filtrado)
        resumo_layout.addStretch()
        
        layout.addWidget(self.frame_resumo_producao_filtrada)
        
        # Carregar dados iniciais
        QTimer.singleShot(100, self.carregar_dados_iniciais_producao_filtrada)
        
        return tab
    
    def carregar_dados_iniciais_producao_filtrada(self):
        """Carrega dados iniciais da aba produção com filtros"""
        # Carregar opções para os combobox
        self.carregar_opcoes_filtros_producao()
        # Carregar dados
        self.carregar_dados_producao_com_filtros()
    
    def carregar_opcoes_filtros_producao(self):
        """Carrega as opções disponíveis para os filtros"""
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Carregar motoristas únicos
            cursor.execute("SELECT DISTINCT motorista FROM pesagens WHERE motorista IS NOT NULL AND motorista != '' ORDER BY motorista")
            motoristas = cursor.fetchall()
            self.combo_motorista_producao.clear()
            self.combo_motorista_producao.addItem("Todos")
            for motorista in motoristas:
                self.combo_motorista_producao.addItem(motorista[0])
            
            # Carregar placas únicas
            cursor.execute("SELECT DISTINCT placa_veiculo FROM pesagens WHERE placa_veiculo IS NOT NULL AND placa_veiculo != '' ORDER BY placa_veiculo")
            placas = cursor.fetchall()
            self.combo_placa_producao.clear()
            self.combo_placa_producao.addItem("Todas")
            for placa in placas:
                self.combo_placa_producao.addItem(placa[0])
            
            # Carregar clientes únicos
            cursor.execute("""
                SELECT DISTINCT c.nome 
                FROM pesagens p 
                LEFT JOIN clientes c ON p.cliente_id = c.id 
                WHERE c.nome IS NOT NULL AND c.nome != '' 
                ORDER BY c.nome
            """)
            clientes = cursor.fetchall()
            self.combo_cliente_producao.clear()
            self.combo_cliente_producao.addItem("Todos")
            for cliente in clientes:
                self.combo_cliente_producao.addItem(cliente[0])
            
            # Carregar traços únicos
            cursor.execute("""
                SELECT DISTINCT t.nome 
                FROM pesagens p 
                LEFT JOIN tracos t ON p.traco_id = t.id 
                WHERE t.nome IS NOT NULL AND t.nome != '' 
                ORDER BY t.nome
            """)
            tracos = cursor.fetchall()
            self.combo_traco_producao.clear()
            self.combo_traco_producao.addItem("Todos")
            for traco in tracos:
                self.combo_traco_producao.addItem(traco[0])
            
            conn.close()
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao carregar opções de filtros: {str(e)}")
    
    def carregar_dados_producao_com_filtros(self):
        """Carrega dados de produção com todos os filtros aplicados"""
        try:
            # Obter valores dos filtros
            data_inicio = self.date_inicial_producao.date().toString("yyyy-MM-dd")
            data_fim = self.date_final_producao.date().toString("yyyy-MM-dd")
            status = self.combo_status_producao.currentText()
            motorista = self.combo_motorista_producao.currentText()
            placa = self.combo_placa_producao.currentText()
            cliente = self.combo_cliente_producao.currentText()
            traco = self.combo_traco_producao.currentText()
            observacao = self.edit_observacao_producao.text().strip()
            volume_min_str = self.spin_volume_min.text().strip()
            
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Construir query dinamicamente
            query = '''
                SELECT 
                    p.id,
                    p.data_pesagem,
                    p.quantidade,
                    t.nome as traco_nome,
                    c.nome as cliente_nome,
                    p.status,
                    p.motorista,
                    p.placa_veiculo,
                    p.observacao
                FROM pesagens p
                LEFT JOIN tracos t ON p.traco_id = t.id
                LEFT JOIN clientes c ON p.cliente_id = c.id
                WHERE 1=1
            '''
            
            params = []
            
            # Adicionar filtro de data
            query += " AND date(p.data_pesagem) BETWEEN ? AND ?"
            params.extend([data_inicio, data_fim])
            
            # Filtro de status
            if status != "Todos":
                query += " AND p.status = ?"
                params.append(status)
            
            # Filtro de motorista
            if motorista != "Todos":
                query += " AND p.motorista = ?"
                params.append(motorista)
            
            # Filtro de placa
            if placa != "Todas":
                query += " AND p.placa_veiculo = ?"
                params.append(placa)
            
            # Filtro de cliente
            if cliente != "Todos":
                query += " AND c.nome = ?"
                params.append(cliente)
            
            # Filtro de traço
            if traco != "Todos":
                query += " AND t.nome = ?"
                params.append(traco)
            
            # Filtro de observação
            if observacao:
                query += " AND p.observacao LIKE ?"
                params.append(f'%{observacao}%')
            
            # Filtro de volume mínimo
            if volume_min_str:
                try:
                    volume_min = float(volume_min_str)
                    query += " AND p.quantidade >= ?"
                    params.append(volume_min)
                except ValueError:
                    pass
            
            query += " ORDER BY p.data_pesagem DESC"
            
            cursor.execute(query, params)
            resultados = cursor.fetchall()
            conn.close()
            
            # Preencher tabela
            self.tabela_producao_filtrada.setRowCount(len(resultados))
            
            total_volume = 0
            total_pesagens = len(resultados)
            
            for i, row in enumerate(resultados):
                pesagem_id = row[0]
                data_pesagem_str = row[1]
                quantidade = row[2] or 0
                traco_nome = row[3] or "N/A"
                cliente_nome = row[4] or "N/A"
                status = row[5]
                motorista = row[6] or ""
                placa = row[7] or ""
                observacao = row[8] or ""
                
                # Formatar data
                data_pesagem = self.formatar_data_banco(data_pesagem_str)
                data_formatada = data_pesagem.strftime("%d/%m/%Y")
                hora_formatada = data_pesagem.strftime("%H:%M")
                
                # Determinar cor do status
                if status == "CONCLUÍDO":
                    status_text = "CARREGADO"
                    cor_status = "#4CAF50"
                elif status == "PENDENTE":
                    status_text = "PENDENTE"
                    cor_status = "#FF9800"
                elif status == "CANCELADO":
                    status_text = "CANCELADO"
                    cor_status = "#F44336"
                else:
                    status_text = str(status)
                    cor_status = "#9E9E9E"
                
                # Adicionar à tabela
                self.tabela_producao_filtrada.setItem(i, 0, QTableWidgetItem(data_formatada))
                self.tabela_producao_filtrada.setItem(i, 1, QTableWidgetItem(f"{quantidade:,.2f}"))
                self.tabela_producao_filtrada.setItem(i, 2, QTableWidgetItem(f"PES-{pesagem_id:04d}"))
                self.tabela_producao_filtrada.setItem(i, 3, QTableWidgetItem(traco_nome))
                self.tabela_producao_filtrada.setItem(i, 4, QTableWidgetItem(cliente_nome))
                self.tabela_producao_filtrada.setItem(i, 5, QTableWidgetItem(f"{quantidade:,.2f}"))
                
                status_item = QTableWidgetItem(status_text)
                status_item.setForeground(QColor(cor_status))
                self.tabela_producao_filtrada.setItem(i, 6, status_item)
                
                self.tabela_producao_filtrada.setItem(i, 7, QTableWidgetItem(hora_formatada))
                self.tabela_producao_filtrada.setItem(i, 8, QTableWidgetItem(motorista))
                self.tabela_producao_filtrada.setItem(i, 9, QTableWidgetItem(placa))
                self.tabela_producao_filtrada.setItem(i, 10, QTableWidgetItem(observacao))
                
                total_volume += quantidade
            
            # Atualizar resumo
            self.atualizar_resumo_producao_filtrada(total_volume, total_pesagens)
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao carregar dados de produção: {str(e)}")
    
    def atualizar_resumo_producao_filtrada(self, total_volume, total_pesagens):
        """Atualiza o resumo da produção filtrada"""
        try:
            # Calcular M³ médio
            m3_medio = total_volume / total_pesagens if total_pesagens > 0 else 0
            
            # Determinar status geral
            if total_pesagens == 0:
                status = "Sem dados"
                cor_status = "#9E9E9E"
            elif total_volume > 100:
                status = "Alta Produção"
                cor_status = "#4CAF50"
            elif total_volume > 50:
                status = "Produção Média"
                cor_status = "#FF9800"
            else:
                status = "Baixa Produção"
                cor_status = "#F44336"
            
            # Atualizar labels
            self.label_total_volume_filtrado.setText(f"📦 Volume Total: {total_volume:,.2f} m³")
            self.label_total_pesagens_filtrado.setText(f"⚖️ Total de Pesagens: {total_pesagens}")
            self.label_media_m3_filtrado.setText(f"📊 M³ Médio: {m3_medio:,.2f}")
            
            self.label_status_geral_filtrado.setText(f"📈 Status: {status}")
            self.label_status_geral_filtrado.setStyleSheet(f"color: {cor_status}; font-weight: bold;")
            
        except Exception as e:
            print(f"Erro ao atualizar resumo: {e}")
    
    def limpar_filtros_producao_avancados(self):
        """Limpa todos os filtros da aba produção avançada"""
        self.date_inicial_producao.setDate(QDate.currentDate().addDays(-7))
        self.date_final_producao.setDate(QDate.currentDate())
        self.combo_status_producao.setCurrentIndex(0)
        self.combo_motorista_producao.setCurrentIndex(0)
        self.combo_placa_producao.setCurrentIndex(0)
        self.combo_cliente_producao.setCurrentIndex(0)
        self.combo_traco_producao.setCurrentIndex(0)
        self.edit_observacao_producao.clear()
        self.spin_volume_min.clear()
        self.carregar_dados_producao_com_filtros()
    
    def criar_aba_semanal(self):
        """Cria aba semanal COM FILTROS DIA, MÊS, ANO"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Título da aba
        titulo_semanal = QLabel("📊 RELATÓRIO SEMANAL - USINA BETTO MIX")
        titulo_semanal.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        titulo_semanal.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo_semanal.setStyleSheet("""
            color: white;
            background-color: #0D47A1;
            padding: 10px;
            border-radius: 5px;
        """)
        layout.addWidget(titulo_semanal)
        
        # Frame de controles - FILTROS DIA, MÊS, ANO
        frame_filtros = QGroupBox("🔍 FILTROS DE CONSULTA SEMANAL")
        frame_filtros.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                margin-top: 5px;
                padding-top: 10px;
                background-color: #E3F2FD;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        filtros_layout = QGridLayout(frame_filtros)
        
        # Filtro por Dia
        filtros_layout.addWidget(QLabel("Dia:"), 0, 0)
        self.combo_semanal_dia = QComboBox()
        self.combo_semanal_dia.addItem("Todos os dias")
        for dia in range(1, 32):
            self.combo_semanal_dia.addItem(f"{dia:02d}")
        self.combo_semanal_dia.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_semanal_dia, 0, 1)
        
        # Filtro por Mês
        filtros_layout.addWidget(QLabel("Mês:"), 0, 2)
        self.combo_semanal_mes = QComboBox()
        self.combo_semanal_mes.addItem("Todos os meses")
        meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        for i, mes in enumerate(meses, 1):
            self.combo_semanal_mes.addItem(mes, i)
        self.combo_semanal_mes.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_semanal_mes, 0, 3)
        
        # Filtro por Ano
        filtros_layout.addWidget(QLabel("Ano:"), 0, 4)
        self.combo_semanal_ano = QComboBox()
        self.combo_semanal_ano.addItem("Todos os anos")
        ano_atual = datetime.now().year
        for ano in range(2020, 2826):
            self.combo_semanal_ano.addItem(str(ano))
        self.combo_semanal_ano.setCurrentText(str(ano_atual))
        self.combo_semanal_ano.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_semanal_ano, 0, 5)
        
        # Botão para gerar relatório
        btn_gerar_semanal = QPushButton("📊 Gerar Relatório Semanal")
        btn_gerar_semanal.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_gerar_semanal.clicked.connect(self.gerar_relatorio_semanal)
        filtros_layout.addWidget(btn_gerar_semanal, 0, 6)
        
        # Botão para limpar filtros
        btn_limpar_semanal = QPushButton("🗑️ Limpar")
        btn_limpar_semanal.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_limpar_semanal.clicked.connect(self.limpar_filtros_semanal)
        filtros_layout.addWidget(btn_limpar_semanal, 0, 7)
        
        layout.addWidget(frame_filtros)
        
        # Resumo da semana
        self.frame_resumo_semanal = QFrame()
        self.frame_resumo_semanal.setFrameStyle(QFrame.Shape.StyledPanel)
        self.frame_resumo_semanal.setStyleSheet("""
            background-color: #E3F2FD;
            border: 2px solid #0D47A1;
            border-radius: 5px;
            padding: 10px;
            margin-bottom: 10px;
        """)
        
        resumo_layout = QHBoxLayout(self.frame_resumo_semanal)
        
        self.label_semana_info = QLabel("Semana: Selecione os filtros para visualizar a semana")
        self.label_semana_info.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        
        self.label_total_semanal = QLabel("📦 Total da Semana: 0.00 m³")
        self.label_total_semanal.setFont(QFont("Arial", 11))
        
        self.label_clientes_semanal = QLabel("👥 Clientes: 0")
        self.label_clientes_semanal.setFont(QFont("Arial", 11))
        
        self.label_faturamento_semanal = QLabel("💰 Faturamento: R$ 0,00")
        self.label_faturamento_semanal.setFont(QFont("Arial", 11))
        
        resumo_layout.addWidget(self.label_semana_info)
        resumo_layout.addWidget(self.label_total_semanal)
        resumo_layout.addWidget(self.label_clientes_semanal)
        resumo_layout.addWidget(self.label_faturamento_semanal)
        resumo_layout.addStretch()
        
        layout.addWidget(self.frame_resumo_semanal)
        
        # Tabela semanal
        self.tabela_semanal = QTableWidget()
        self.tabela_semanal.setColumnCount(8)
        self.tabela_semanal.setHorizontalHeaderLabels([
            "Dia", "Produção (m³)", "Clientes", "Consumo Total", 
            "Cimento (kg)", "Brita (kg)", "Areia (kg)", "Faturamento"
        ])
        
        # Configurar largura das colunas
        header = self.tabela_semanal.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
        
        # Aplicar estilo da tabela com paleta de azuis
        self.tabela_semanal.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #E3F2FD;
                selection-background-color: #BBDEFB;
                border: 1px solid #1976D2;
                gridline-color: #BBDEFB;
            }
            QHeaderView::section {
                background-color: #0D47A1;
                color: white;
                padding: 5px;
                border: 1px solid #0D47A1;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        layout.addWidget(self.tabela_semanal)
        
        # Conectar os filtros para atualização automática (SINCRONIZAÇÃO)
        self.combo_semanal_dia.currentIndexChanged.connect(self.gerar_relatorio_semanal)
        self.combo_semanal_mes.currentIndexChanged.connect(self.gerar_relatorio_semanal)
        self.combo_semanal_ano.currentIndexChanged.connect(self.gerar_relatorio_semanal)
        
        # Carregar dados iniciais
        QTimer.singleShot(100, self.carregar_dados_iniciais_semanal)
        
        return tab
    
    def carregar_dados_iniciais_semanal(self):
        """Carrega dados iniciais da aba semanal"""
        hoje = datetime.now()
        self.combo_semanal_dia.setCurrentText(f"{hoje.day:02d}")
        self.combo_semanal_mes.setCurrentIndex(hoje.month)
        self.combo_semanal_ano.setCurrentText(str(hoje.year))
        self.gerar_relatorio_semanal()
    
    def limpar_filtros_semanal(self):
        """Limpa todos os filtros da aba semanal"""
        self.combo_semanal_dia.setCurrentIndex(0)
        self.combo_semanal_mes.setCurrentIndex(0)
        self.combo_semanal_ano.setCurrentIndex(0)
        self.gerar_relatorio_semanal()
    
    def criar_aba_mensal(self):
        """Cria aba mensal COMPLETA com todos os filtros solicitados"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Título da aba
        titulo_mensal = QLabel("📊 RELATÓRIO MENSAL - USINA BETTO MIX")
        titulo_mensal.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        titulo_mensal.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo_mensal.setStyleSheet("""
            color: white;
            background-color: #0D47A1;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 10px;
        """)
        layout.addWidget(titulo_mensal)
        
        # Frame de controles - FILTROS COMPLETOS
        frame_filtros = QGroupBox("🔍 FILTROS DE CONSULTA MENSAL")
        frame_filtros.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                margin-top: 5px;
                padding-top: 10px;
                background-color: #E3F2FD;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        filtros_layout = QGridLayout(frame_filtros)
        
        # Filtro 1: Data Específica (Calendário)
        filtros_layout.addWidget(QLabel("Data Específica:"), 0, 0)
        self.date_especifica_mensal = QDateEdit()
        self.date_especifica_mensal.setDisplayFormat("dd/MM/yyyy")
        self.date_especifica_mensal.setDate(QDate.currentDate())
        self.date_especifica_mensal.setCalendarPopup(True)
        self.date_especifica_mensal.setStyleSheet("""
            QDateEdit {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QDateEdit:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.date_especifica_mensal, 0, 1)
        
        # Filtro 2: Dia do Mês (1-31)
        filtros_layout.addWidget(QLabel("Dia do Mês:"), 0, 2)
        self.combo_mensal_dia = QComboBox()
        self.combo_mensal_dia.addItem("Todos os dias")
        for dia in range(1, 32):
            self.combo_mensal_dia.addItem(f"{dia}")
        self.combo_mensal_dia.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_mensal_dia, 0, 3)
        
        # Filtro 3: Número da Semana (1-53)
        filtros_layout.addWidget(QLabel("Número da Semana:"), 0, 4)
        self.combo_mensal_semana = QComboBox()
        self.combo_mensal_semana.addItem("Todas as semanas")
        for semana in range(1, 54):
            self.combo_mensal_semana.addItem(f"{semana}")
        self.combo_mensal_semana.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_mensal_semana, 0, 5)
        
        # Filtro 4: Mês (Janeiro-Dezembro)
        filtros_layout.addWidget(QLabel("Mês:"), 1, 0)
        self.combo_mensal_mes = QComboBox()
        self.combo_mensal_mes.addItem("Todos os meses")
        meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        for i, mes in enumerate(meses, 1):
            self.combo_mensal_mes.addItem(mes, i)
        self.combo_mensal_mes.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_mensal_mes, 1, 1)
        
        # Filtro 5: Ano (2020-2826)
        filtros_layout.addWidget(QLabel("Ano:"), 1, 2)
        self.combo_mensal_ano = QComboBox()
        self.combo_mensal_ano.addItem("Todos os anos")
        ano_atual = datetime.now().year
        for ano in range(2020, 2827):
            self.combo_mensal_ano.addItem(str(ano))
        self.combo_mensal_ano.setCurrentText(str(ano_atual))
        self.combo_mensal_ano.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_mensal_ano, 1, 3)
        
        # Botão para gerar relatório
        btn_gerar_mensal = QPushButton("📊 Gerar Relatório Mensal")
        btn_gerar_mensal.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_gerar_mensal.clicked.connect(self.gerar_relatorio_mensal)
        filtros_layout.addWidget(btn_gerar_mensal, 1, 4)
        
        # Botão para limpar filtros
        btn_limpar_mensal = QPushButton("🗑️ Limpar Filtros")
        btn_limpar_mensal.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_limpar_mensal.clicked.connect(self.limpar_filtros_mensal)
        filtros_layout.addWidget(btn_limpar_mensal, 1, 5)
        
        layout.addWidget(frame_filtros)
        
        # Resumo do período
        self.frame_resumo_mensal = QFrame()
        self.frame_resumo_mensal.setFrameStyle(QFrame.Shape.StyledPanel)
        self.frame_resumo_mensal.setStyleSheet("""
            background-color: #E3F2FD;
            border: 2px solid #0D47A1;
            border-radius: 5px;
            padding: 10px;
            margin-bottom: 10px;
        """)
        
        resumo_layout = QHBoxLayout(self.frame_resumo_mensal)
        
        self.label_periodo_info = QLabel("Período: Selecione os filtros")
        self.label_periodo_info.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        
        self.label_total_mensal = QLabel("📦 Volume Total: 0.00 m³")
        self.label_total_mensal.setFont(QFont("Arial", 11))
        
        self.label_clientes_mensal = QLabel("👥 Clientes: 0")
        self.label_clientes_mensal.setFont(QFont("Arial", 11))
        
        self.label_faturamento_mensal = QLabel("💰 Faturamento: R$ 0,00")
        self.label_faturamento_mensal.setFont(QFont("Arial", 11))
        
        resumo_layout.addWidget(self.label_periodo_info)
        resumo_layout.addWidget(self.label_total_mensal)
        resumo_layout.addWidget(self.label_clientes_mensal)
        resumo_layout.addWidget(self.label_faturamento_mensal)
        resumo_layout.addStretch()
        
        layout.addWidget(self.frame_resumo_mensal)
        
        # Tabela mensal com 10 colunas
        self.tabela_mensal = QTableWidget()
        self.tabela_mensal.setColumnCount(10)
        self.tabela_mensal.setHorizontalHeaderLabels([
            "Data", "Dia", "Semana", "Mês", "Ano", 
            "Produção (m³)", "Clientes", "Consumo Total (kg)", 
            "Custo Total (R$)", "Faturamento (R$)"
        ])
        
        # Configurar largura das colunas
        header = self.tabela_mensal.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)
        
        # Aplicar estilo da tabela com paleta de azuis
        self.tabela_mensal.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #E3F2FD;
                selection-background-color: #BBDEFB;
                border: 1px solid #1976D2;
                gridline-color: #BBDEFB;
            }
            QHeaderView::section {
                background-color: #0D47A1;
                color: white;
                padding: 5px;
                border: 1px solid #0D47A1;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        layout.addWidget(self.tabela_mensal)
        
        # Conectar os filtros para atualização automática (SINCRONIZAÇÃO)
        self.date_especifica_mensal.dateChanged.connect(self.gerar_relatorio_mensal)
        self.combo_mensal_dia.currentIndexChanged.connect(self.gerar_relatorio_mensal)
        self.combo_mensal_semana.currentIndexChanged.connect(self.gerar_relatorio_mensal)
        self.combo_mensal_mes.currentIndexChanged.connect(self.gerar_relatorio_mensal)
        self.combo_mensal_ano.currentIndexChanged.connect(self.gerar_relatorio_mensal)
        
        # Carregar dados iniciais
        QTimer.singleShot(100, self.carregar_dados_iniciais_mensal)
        
        return tab
    
    def carregar_dados_iniciais_mensal(self):
        """Carrega dados iniciais da aba mensal"""
        hoje = datetime.now()
        self.date_especifica_mensal.setDate(QDate(hoje.year, hoje.month, hoje.day))
        self.combo_mensal_dia.setCurrentText(f"{hoje.day}")
        self.combo_mensal_mes.setCurrentIndex(hoje.month)
        self.combo_mensal_ano.setCurrentText(str(hoje.year))
        self.gerar_relatorio_mensal()
    
    def limpar_filtros_mensal(self):
        """Limpa todos os filtros da aba mensal"""
        self.date_especifica_mensal.setDate(QDate.currentDate())
        self.combo_mensal_dia.setCurrentIndex(0)
        self.combo_mensal_semana.setCurrentIndex(0)
        self.combo_mensal_mes.setCurrentIndex(0)
        self.combo_mensal_ano.setCurrentIndex(0)
        self.gerar_relatorio_mensal()
    
    def criar_aba_gastos(self):
        """Cria aba de gastos COMPLETA com data, dia, semana, mês, ano até 2826"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Título da aba
        titulo_gastos = QLabel("💰 RELATÓRIO DE GASTOS - USINA BETTO MIX")
        titulo_gastos.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        titulo_gastos.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo_gastos.setStyleSheet("""
            color: white;
            background-color: #0D47A1;
            padding: 10px;
            border-radius: 5px;
            margin-bottom: 10px;
        """)
        layout.addWidget(titulo_gastos)
        
        # Frame de controles - FILTROS COMPLETOS
        frame_filtros = QGroupBox("🔍 FILTROS DE CONSULTA DE GASTOS")
        frame_filtros.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                margin-top: 5px;
                padding-top: 10px;
                background-color: #E3F2FD;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        
        filtros_layout = QGridLayout(frame_filtros)
        
        # Filtro 1: Data Específica (Calendário)
        filtros_layout.addWidget(QLabel("Data Específica:"), 0, 0)
        self.date_especifica_gastos = QDateEdit()
        self.date_especifica_gastos.setDisplayFormat("dd/MM/yyyy")
        self.date_especifica_gastos.setDate(QDate.currentDate())
        self.date_especifica_gastos.setCalendarPopup(True)
        self.date_especifica_gastos.setMaximumDate(QDate(2826, 12, 31))
        self.date_especifica_gastos.setStyleSheet("""
            QDateEdit {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QDateEdit:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.date_especifica_gastos, 0, 1)
        
        # Filtro 2: Dia do Mês (1-31)
        filtros_layout.addWidget(QLabel("Dia do Mês:"), 0, 2)
        self.combo_gastos_dia = QComboBox()
        self.combo_gastos_dia.addItem("Todos os dias")
        for dia in range(1, 32):
            self.combo_gastos_dia.addItem(f"{dia:02d}")
        self.combo_gastos_dia.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_gastos_dia, 0, 3)
        
        # Filtro 3: Número da Semana (1-53)
        filtros_layout.addWidget(QLabel("Número da Semana:"), 0, 4)
        self.combo_gastos_semana = QComboBox()
        self.combo_gastos_semana.addItem("Todas as semanas")
        for semana in range(1, 54):
            self.combo_gastos_semana.addItem(f"{semana}")
        self.combo_gastos_semana.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_gastos_semana, 0, 5)
        
        # Filtro 4: Mês (Janeiro-Dezembro)
        filtros_layout.addWidget(QLabel("Mês:"), 1, 0)
        self.combo_gastos_mes = QComboBox()
        self.combo_gastos_mes.addItem("Todos os meses")
        meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        for i, mes in enumerate(meses, 1):
            self.combo_gastos_mes.addItem(mes, i)
        self.combo_gastos_mes.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_gastos_mes, 1, 1)
        
        # Filtro 5: Ano (2020-2826)
        filtros_layout.addWidget(QLabel("Ano:"), 1, 2)
        self.combo_gastos_ano = QComboBox()
        self.combo_gastos_ano.addItem("Todos os anos")
        ano_atual = datetime.now().year
        for ano in range(2020, 2827):
            self.combo_gastos_ano.addItem(str(ano))
        self.combo_gastos_ano.setCurrentText(str(ano_atual))
        self.combo_gastos_ano.setStyleSheet("""
            QComboBox {
                border: 1px solid #2196F3;
                border-radius: 3px;
                padding: 3px;
                background-color: white;
            }
            QComboBox:focus {
                border: 2px solid #1976D2;
            }
        """)
        filtros_layout.addWidget(self.combo_gastos_ano, 1, 3)
        
        # Botão para gerar relatório
        btn_gerar_gastos = QPushButton("📊 Gerar Relatório de Gastos")
        btn_gerar_gastos.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_gerar_gastos.clicked.connect(self.gerar_relatorio_gastos_corrigido)
        filtros_layout.addWidget(btn_gerar_gastos, 1, 4)
        
        # Botão para limpar filtros
        btn_limpar_gastos = QPushButton("🗑️ Limpar Filtros")
        btn_limpar_gastos.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_limpar_gastos.clicked.connect(self.limpar_filtros_gastos)
        filtros_layout.addWidget(btn_limpar_gastos, 1, 5)
        
        # Botão para diagnóstico
        btn_diagnostico = QPushButton("🔧 Diagnóstico")
        btn_diagnostico.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_diagnostico.clicked.connect(self.testar_conexao_gastos)
        filtros_layout.addWidget(btn_diagnostico, 1, 6)
        
        layout.addWidget(frame_filtros)
        
        # Tabela de gastos - AUMENTADA PARA 11 COLUNAS
        self.tabela_gastos = QTableWidget()
        self.tabela_gastos.setColumnCount(11)
        self.tabela_gastos.setHorizontalHeaderLabels([
            "Data", "Dia", "Semana", "Mês", "Ano", 
            "Fornecedor", "Material", "Quantidade", 
            "Unidade", "Valor Unitário (R$)", "Valor Total (R$)"
        ])
        
        # Configurar largura das colunas
        header = self.tabela_gastos.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(10, QHeaderView.ResizeMode.ResizeToContents)
        
        # Aplicar estilo da tabela com paleta de azuis
        self.tabela_gastos.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #E3F2FD;
                selection-background-color: #BBDEFB;
                border: 1px solid #1976D2;
                gridline-color: #BBDEFB;
            }
            QHeaderView::section {
                background-color: #0D47A1;
                color: white;
                padding: 5px;
                border: 1px solid #0D47A1;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        layout.addWidget(self.tabela_gastos)
        
        # Total de gastos
        self.label_total_gastos = QLabel("💰 Total de Gastos: R$ 0,00")
        self.label_total_gastos.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.label_total_gastos.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        self.label_total_gastos.setStyleSheet("""
            background-color: #E3F2FD;
            color: #0D47A1;
            padding: 15px;
            border-radius: 8px;
            border: 2px solid #0D47A1;
            margin-top: 10px;
        """)
        layout.addWidget(self.label_total_gastos)
        
        # Conectar os filtros para atualização automática (SINCRONIZAÇÃO)
        self.date_especifica_gastos.dateChanged.connect(self.gerar_relatorio_gastos_corrigido)
        self.combo_gastos_dia.currentIndexChanged.connect(self.gerar_relatorio_gastos_corrigido)
        self.combo_gastos_semana.currentIndexChanged.connect(self.gerar_relatorio_gastos_corrigido)
        self.combo_gastos_mes.currentIndexChanged.connect(self.gerar_relatorio_gastos_corrigido)
        self.combo_gastos_ano.currentIndexChanged.connect(self.gerar_relatorio_gastos_corrigido)
        
        # Carregar dados iniciais
        QTimer.singleShot(100, self.carregar_dados_iniciais_gastos)
        
        return tab
    
    def carregar_dados_iniciais_gastos(self):
        """Carrega dados iniciais da aba de gastos - AJUSTADA para filtros padrão"""
        hoje = datetime.now()
        
        # CORREÇÃO: Setar filtros para valores mais amplos por padrão
        # Data específica = hoje
        self.date_especifica_gastos.setDate(QDate(hoje.year, hoje.month, hoje.day))
        
        # Dia = "Todos os dias" (índice 0)
        self.combo_gastos_dia.setCurrentIndex(0)
        
        # Semana = "Todas as semanas" (índice 0)
        self.combo_gastos_semana.setCurrentIndex(0)
        
        # Mês = Mês atual (não "Todos os meses")
        self.combo_gastos_mes.setCurrentIndex(hoje.month)
        
        # Ano = Ano atual (não "Todos os anos")
        ano_atual = str(hoje.year)
        index_ano_atual = self.combo_gastos_ano.findText(ano_atual)
        if index_ano_atual >= 0:
            self.combo_gastos_ano.setCurrentIndex(index_ano_atual)
        
        # Pequeno delay para garantir que a UI está pronta
        QTimer.singleShot(500, self.gerar_relatorio_gastos_corrigido)
    
    def limpar_filtros_gastos(self):
        """Limpa todos os filtros da aba de gastos"""
        self.date_especifica_gastos.setDate(QDate.currentDate())
        self.combo_gastos_dia.setCurrentIndex(0)
        self.combo_gastos_semana.setCurrentIndex(0)
        self.combo_gastos_mes.setCurrentIndex(0)
        self.combo_gastos_ano.setCurrentIndex(0)
        self.gerar_relatorio_gastos_corrigido()
    
    def criar_coluna_valor_unitario(self):
        """Cria a coluna valor_unitario na tabela historico_estoque se ela não existir"""
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Verificar se a coluna já existe
            cursor.execute("PRAGMA table_info(historico_estoque)")
            colunas = cursor.fetchall()
            colunas_nomes = [col[1] for col in colunas]
            
            if "valor_unitario" not in colunas_nomes:
                # Adicionar a coluna
                cursor.execute("ALTER TABLE historico_estoque ADD COLUMN valor_unitario REAL DEFAULT 0")
                conn.commit()
                
                # Atualizar registros existentes com valor unitário padrão baseado no material
                cursor.execute('''
                    UPDATE historico_estoque 
                    SET valor_unitario = CASE 
                        WHEN material LIKE '%Cimento%' THEN 0.50
                        WHEN material LIKE '%Brita%' THEN 0.15
                        WHEN material LIKE '%Areia%' THEN 0.10
                        WHEN material LIKE '%Aditivo%' THEN 2.00
                        ELSE 0.10
                    END
                    WHERE tipo = 'ENTRADA' AND (valor_unitario IS NULL OR valor_unitario = 0)
                ''')
                conn.commit()
                
                QMessageBox.information(self, "Coluna Criada", 
                    "Coluna 'valor_unitario' criada com sucesso na tabela historico_estoque!\n"
                    "Valores padrão foram atribuídos aos registros existentes.")
            
            conn.close()
            return True
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao criar coluna: {str(e)}")
            return False
    
    def testar_conexao_gastos(self):
        """Testa a conexão com o banco de dados para gastos"""
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            mensagem_diagnostico = "🔍 DIAGNÓSTICO DO BANCO DE DADOS 🔍\n\n"
            
            # Verificar todas as tabelas
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
            tabelas = cursor.fetchall()
            mensagem_diagnostico += f"📋 TABELAS ENCONTRADAS ({len(tabelas)}):\n"
            for tabela in tabelas:
                mensagem_diagnostico += f"  • {tabela[0]}\n"
            
            # Verificar se a tabela historico_estoque existe
            if any("historico_estoque" in tabela[0] for tabela in tabelas):
                # Verificar estrutura da tabela historico_estoque
                cursor.execute("PRAGMA table_info(historico_estoque)")
                colunas = cursor.fetchall()
                mensagem_diagnostico += f"\n📊 ESTRUTURA DA TABELA 'historico_estoque' ({len(colunas)} colunas):\n"
                for coluna in colunas:
                    mensagem_diagnostico += f"  • {coluna[1]} ({coluna[2]})\n"
                
                # Verificar se a coluna valor_unitario existe
                colunas_nomes = [col[1] for col in colunas]
                if "valor_unitario" in colunas_nomes:
                    mensagem_diagnostico += "✓ Coluna 'valor_unitario' EXISTE\n"
                else:
                    mensagem_diagnostico += "✗ Coluna 'valor_unitario' NÃO EXISTE\n"
                    mensagem_diagnostico += "  • É necessário criar esta coluna para relatórios de gastos\n"
                
                # Contar registros por tipo
                cursor.execute("SELECT tipo, COUNT(*) FROM historico_estoque GROUP BY tipo ORDER BY tipo")
                contagem = cursor.fetchall()
                mensagem_diagnostico += f"\n📈 CONTAGEM DE REGISTROS POR TIPO:\n"
                total_registros = 0
                for tipo, quantidade in contagem:
                    mensagem_diagnostico += f"  • {tipo}: {quantidade} registros\n"
                    total_registros += quantidade
                mensagem_diagnostico += f"  • TOTAL: {total_registros} registros\n"
                
                # Buscar alguns exemplos de ENTRADA
                cursor.execute("""
                    SELECT data_movimentacao, material, quantidade, destino, tipo 
                    FROM historico_estoque 
                    WHERE tipo = 'ENTRADA' 
                    ORDER BY data_movimentacao DESC 
                    LIMIT 5
                """)
                exemplos = cursor.fetchall()
                mensagem_diagnostico += f"\n📅 ÚLTIMAS 5 ENTRADAS:\n"
                if exemplos:
                    for exemplo in exemplos:
                        data, material, quantidade, destino, tipo = exemplo
                        mensagem_diagnostico += f"  • {data} - {material} - {quantidade}kg - {destino or 'Sistema'}\n"
                else:
                    mensagem_diagnostico += "  • Nenhuma entrada encontrada\n"
                
            else:
                mensagem_diagnostico += "\n❌ ERRO: Tabela 'historico_estoque' NÃO ENCONTRADA!\n"
                mensagem_diagnostico += "  • A tabela de histórico de estoque não existe no banco de dados.\n"
                mensagem_diagnostico += "  • Execute o comando SQL para criar a tabela.\n"
                mensagem_diagnostico += "  • Ou verifique se o nome da tabela está correto.\n"
            
            conn.close()
            
            # Mostrar resultado em um QMessageBox
            dialog = QMessageBox(self)
            dialog.setWindowTitle("🔧 Diagnóstico do Banco de Dados")
            dialog.setText(mensagem_diagnostico)
            dialog.setIcon(QMessageBox.Icon.Information)
            
            # Adicionar botão para criar coluna se não existir
            if "historico_estoque" in [t[0] for t in tabelas]:
                cursor.execute("PRAGMA table_info(historico_estoque)")
                colunas = cursor.fetchall()
                colunas_nomes = [col[1] for col in colunas]
                
                if "valor_unitario" not in colunas_nomes:
                    btn_criar_coluna = dialog.addButton("Criar Coluna valor_unitario", QMessageBox.ButtonRole.ActionRole)
                    dialog.addButton(QMessageBox.StandardButton.Ok)
                    
                    dialog.exec()
                    
                    if dialog.clickedButton() == btn_criar_coluna:
                        if self.criar_coluna_valor_unitario():
                            # Recarregar relatório
                            self.gerar_relatorio_gastos_corrigido()
                else:
                    dialog.exec()
            else:
                dialog.exec()
            
        except sqlite3.Error as e:
            QMessageBox.warning(self, "Erro no Diagnóstico", f"❌ ERRO SQLite: {str(e)}\n\nVerifique se o banco de dados existe e está acessível.")
        except Exception as e:
            QMessageBox.warning(self, "Erro no Diagnóstico", f"❌ ERRO GERAL: {str(e)}")
    
    def criar_aba_estoque_detalhado(self):
        """Cria aba de estoque detalhado com todos os materiais"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Título
        titulo_estoque = QLabel("📦 ESTOQUE DETALHADO DE MATERIAIS")
        titulo_estoque.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        titulo_estoque.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo_estoque.setStyleSheet("color: #0D47A1; padding: 10px; border: 1px solid #1976D2; background-color: #E3F2FD; border-radius: 5px;")
        layout.addWidget(titulo_estoque)
        
        # Botão atualizar
        btn_atualizar = QPushButton("🔄 Atualizar Estoque")
        btn_atualizar.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1976D2, stop: 1 #1565C0);
                color: white;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
                border: 1px solid #0D47A1;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                                  stop: 0 #1565C0, stop: 1 #0D47A1);
            }
        """)
        btn_atualizar.clicked.connect(self.carregar_estoque_detalhado)
        layout.addWidget(btn_atualizar)
        
        # Tabela de estoque
        self.tabela_estoque_detalhado = QTableWidget()
        self.tabela_estoque_detalhado.setColumnCount(7)
        self.tabela_estoque_detalhado.setHorizontalHeaderLabels([
            "Material", "Quantidade", "Unidade", "Estoque Mínimo", 
            "Estoque Ideal", "Status", "Última Atualização"
        ])
        
        # Configurar largura das colunas
        header = self.tabela_estoque_detalhado.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        
        # Aplicar estilo da tabela com paleta de azuis
        self.tabela_estoque_detalhado.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #E3F2FD;
                selection-background-color: #BBDEFB;
                border: 1px solid #1976D2;
                gridline-color: #BBDEFB;
            }
            QHeaderView::section {
                background-color: #0D47A1;
                color: white;
                padding: 5px;
                border: 1px solid #0D47A1;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
            }
        """)
        
        layout.addWidget(self.tabela_estoque_detalhado)
        
        # Carregar dados inicialmente
        self.carregar_estoque_detalhado()
        
        return tab
    
    # ========== FUNÇÕES AUXILIARES ==========
    
    def formatar_data_banco(self, data_str):
        """Formata a data do banco para datetime, lidando com diferentes formatos"""
        formatos_possiveis = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y"
        ]
        
        for formato in formatos_possiveis:
            try:
                return datetime.strptime(data_str, formato)
            except ValueError:
                continue
        
        print(f"[AVISO] Formato de data não reconhecido: {data_str}")
        return datetime.now()
    
    def get_cor_material(self, material):
        """Retorna a cor do material para estilização"""
        cores = {
            "Cimento": "#1565C0",
            "Brita 0": "#0D47A1",
            "Brita 1": "#1976D2",
            "Areia Média": "#2196F3",
            "Pó de Brita": "#64B5F6",
            "Aditivo": "#1E88E5",
            "Água": "#42A5F5"
        }
        return cores.get(material, "#000000")
    
    # ========== FUNÇÕES DE SINCRONIZAÇÃO E CONSULTA ==========
    
    def carregar_dados_producao(self):
        """MANTIDA PARA COMPATIBILIDADE - usa a nova função com filtros"""
        self.carregar_dados_producao_com_filtros()
    
    def carregar_dados_iniciais_producao(self):
        """MANTIDA PARA COMPATIBILIDADE"""
        self.carregar_dados_iniciais_producao_filtrada()
    
    def limpar_filtros_producao(self):
        """MANTIDA PARA COMPATIBILIDADE"""
        self.limpar_filtros_producao_avancados()
    
    def carregar_estoque_detalhado(self):
        """Carrega o estoque detalhado com Estoque Ideal fixo em 5000 e remove a linha 'Areia'"""
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT 
                    material, 
                    quantidade, 
                    unidade,
                    estoque_minimo,
                    data_atualizacao
                FROM estoque 
                WHERE material != 'Areia'
                ORDER BY 
                    CASE material 
                        WHEN 'Cimento' THEN 1
                        WHEN 'Brita 0' THEN 2
                        WHEN 'Brita 1' THEN 3
                        WHEN 'Areia Média' THEN 4
                        WHEN 'Pó de Brita' THEN 5
                        WHEN 'Aditivo' THEN 6
                        WHEN 'Água' THEN 7
                        ELSE 8
                    END
            ''')
            
            estoque = cursor.fetchall()
            conn.close()
            
            self.tabela_estoque_detalhado.setRowCount(len(estoque))
            
            for i, item in enumerate(estoque):
                material, quantidade, unidade, estoque_minimo, data_atualizacao = item
                
                # CORREÇÃO: Estoque Ideal fixo em 5000 para todos os materiais
                estoque_ideal = 5000
                
                # Determinar status baseado no estoque atual vs 5000
                if quantidade <= 0:
                    status = "ESGOTADO"
                    cor_status = "#F44336"
                elif quantidade < estoque_minimo:
                    status = "CRÍTICO"
                    cor_status = "#FF9800"
                elif quantidade < estoque_ideal:
                    status = "ATENÇÃO"
                    cor_status = "#FFC107"
                else:
                    status = "OK"
                    cor_status = "#4CAF50"
                
                # Adicionar itens na tabela
                self.tabela_estoque_detalhado.setItem(i, 0, QTableWidgetItem(material))
                self.tabela_estoque_detalhado.setItem(i, 1, QTableWidgetItem(f"{quantidade:,.2f}"))
                self.tabela_estoque_detalhado.setItem(i, 2, QTableWidgetItem(unidade))
                self.tabela_estoque_detalhado.setItem(i, 3, QTableWidgetItem(f"{estoque_minimo:,.2f}"))
                self.tabela_estoque_detalhado.setItem(i, 4, QTableWidgetItem(f"{estoque_ideal:,.2f}"))
                
                status_item = QTableWidgetItem(status)
                status_item.setForeground(QColor(cor_status))
                self.tabela_estoque_detalhado.setItem(i, 5, status_item)
                
                self.tabela_estoque_detalhado.setItem(i, 6, QTableWidgetItem(data_atualizacao or "-"))
                
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao carregar estoque: {str(e)}")
    
    def gerar_relatorios(self):
        """Gera todos os relatórios"""
        try:
            self.gerar_relatorio_diario()
            self.carregar_dados_producao_com_filtros()  # AGORA COM FILTROS COMPLETOS
            self.gerar_relatorio_semanal()
            self.gerar_relatorio_mensal()
            self.gerar_relatorio_gastos_corrigido()
            self.carregar_estoque_detalhado()
            
            QMessageBox.information(self, "✅ Sucesso", "Todos os relatórios foram atualizados e sincronizados!")
            
        except Exception as e:
            QMessageBox.warning(self, "⚠️ Aviso", f"Erro ao gerar relatórios: {str(e)}")
    
    def obter_consumo_material_dia(self, material, data):
        """Obtém o consumo de um material em uma data específica - CORRIGIDO PARA ÁGUA"""
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Primeiro, verifica se o material é água
            if material == "Água":
                # Para água, precisamos consultar as pesagens e extrair do JSON
                cursor.execute('''
                    SELECT materiais_json 
                    FROM pesagens 
                    WHERE status = 'CONCLUÍDO' 
                    AND date(data_pesagem) = ?
                ''', (data,))
                
                resultados = cursor.fetchall()
                consumo_total = 0
                
                for resultado in resultados:
                    if resultado[0]:
                        try:
                            materiais_json = json.loads(resultado[0])
                            if "Água" in materiais_json:
                                quantidade_litros = materiais_json["Água"].get("quantidade", 0)
                                consumo_total += quantidade_litros
                        except:
                            continue
                
                conn.close()
                return consumo_total
            else:
                # Para outros materiais, usa a consulta normal
                cursor.execute('''
                    SELECT SUM(quantidade) 
                    FROM historico_estoque 
                    WHERE material = ? 
                    AND tipo = 'SAIDA' 
                    AND date(data_movimentacao) = ?
                    AND destino LIKE '%Pesagem%'
                ''', (material, data))
                
                resultado = cursor.fetchone()
                consumo = resultado[0] if resultado[0] else 0
                
                conn.close()
                return consumo
                
        except Exception as e:
            print(f"Erro em obter_consumo_material_dia para {material}: {e}")
            return 0
    
    def obter_consumo_material_periodo(self, material, data_inicio, data_fim):
        """Obtém o consumo de um material em um período específico - CORRIGIDO PARA ÁGUA"""
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Primeiro, verifica se o material é água
            if material == "Água":
                # Para água, precisamos consultar as pesagens e extrair do JSON
                cursor.execute('''
                    SELECT materiais_json 
                    FROM pesagens 
                    WHERE status = 'CONCLUÍDO' 
                    AND date(data_pesagem) BETWEEN ? AND ?
                ''', (data_inicio, data_fim))
                
                resultados = cursor.fetchall()
                consumo_total = 0
                
                for resultado in resultados:
                    if resultado[0]:
                        try:
                            materiais_json = json.loads(resultado[0])
                            if "Água" in materiais_json:
                                quantidade_litros = materiais_json["Água"].get("quantidade", 0)
                                consumo_total += quantidade_litros
                        except:
                            continue
                
                conn.close()
                return consumo_total
            else:
                # Para outros materiais, usa a consulta normal
                cursor.execute('''
                    SELECT SUM(quantidade) 
                    FROM historico_estoque 
                    WHERE material = ? 
                    AND tipo = 'SAIDA' 
                    AND date(data_movimentacao) BETWEEN ? AND ?
                    AND destino LIKE '%Pesagem%'
                ''', (material, data_inicio, data_fim))
                
                resultado = cursor.fetchone()
                consumo = resultado[0] if resultado[0] else 0
                
                conn.close()
                return consumo
                
        except Exception as e:
            print(f"Erro em obter_consumo_material_periodo para {material}: {e}")
            return 0
    
    def obter_estoque_inicial_dia(self, material, data):
        """Obtém o estoque inicial de um material em uma data"""
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT quantidade FROM estoque WHERE material = ?
            ''', (material,))
            
            estoque_atual = cursor.fetchone()
            estoque_atual = estoque_atual[0] if estoque_atual else 0
            
            consumo = self.obter_consumo_material_dia(material, data)
            
            estoque_inicial = estoque_atual + consumo
            
            conn.close()
            return estoque_inicial
            
        except:
            return 0
    
    def obter_estoque_inicial_periodo(self, material, data_inicio):
        """Obtém o estoque inicial de um material no início do período"""
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Buscar estoque atual
            cursor.execute('''
                SELECT quantidade FROM estoque WHERE material = ?
            ''', (material,))
            
            estoque_atual = cursor.fetchone()
            estoque_atual = estoque_atual[0] if estoque_atual else 0
            
            # Buscar consumo desde o início do período até hoje
            hoje = datetime.now().strftime("%Y-%m-%d")
            consumo = self.obter_consumo_material_periodo(material, data_inicio, hoje)
            
            # Estoque inicial = estoque atual + consumo do período
            estoque_inicial = estoque_atual + consumo
            
            conn.close()
            return estoque_inicial
            
        except:
            return 0
    
    def gerar_relatorio_diario(self):
        """Gera relatório diário com filtros por dia, mês e ano"""
        try:
            # Obter filtros
            filtro_dia = self.combo_dia_filtro.currentText()
            filtro_mes = self.combo_mes_filtro.currentData()
            filtro_ano = self.combo_ano_filtro.currentText()
            
            # Construir período de filtro
            periodo_str = "Período: "
            conditions = []
            params = []
            
            if filtro_dia != "Todos os dias":
                conditions.append("CAST(strftime('%d', data_pesagem) AS INTEGER) = ?")
                params.append(int(filtro_dia))
                periodo_str += f"Dia {filtro_dia}, "
            
            if filtro_mes:
                conditions.append("CAST(strftime('%m', data_pesagem) AS INTEGER) = ?")
                params.append(filtro_mes)
                meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
                periodo_str += f"{meses[filtro_mes-1]}, "
            
            if filtro_ano != "Todos os anos":
                conditions.append("strftime('%Y', data_pesagem) = ?")
                params.append(filtro_ano)
                periodo_str += f"Ano {filtro_ano}"
            
            # Remover vírgula extra no final
            if periodo_str.endswith(", "):
                periodo_str = periodo_str[:-2]
            
            if periodo_str == "Período: ":
                periodo_str = "Período: Todos os dias"
            
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # ========== 1. VOLUME M³ CARREGADO NO PERÍODO ==========
            query_base = '''
                SELECT 
                    SUM(quantidade) as total_volume,
                    COUNT(*) as num_pesagens,
                    COUNT(DISTINCT traco_id) as num_tracos,
                    COUNT(DISTINCT cliente_id) as num_clientes
                FROM pesagens 
                WHERE status = 'CONCLUÍDO'
            '''
            
            if conditions:
                query_base += " AND " + " AND ".join(conditions)
            
            cursor.execute(query_base, params)
            resultado = cursor.fetchone()
            volume_total = resultado[0] or 0
            num_pesagens = resultado[1] or 0
            num_tracos = resultado[2] or 0
            num_clientes = resultado[3] or 0
            
            # Atualizar labels de produção
            self.label_volume_total.setText(f"📦 Volume total carregado: {volume_total:,.2f} m³")
            self.label_num_pesagens.setText(f"⚖️ Número de pesagens: {num_pesagens}")
            self.label_tracos_utilizados.setText(f"🏗️ Traços utilizados: {num_tracos}")
            self.label_clientes_atendidos.setText(f"👥 Clientes atendidos: {num_clientes}")
            
            # ========== 2. SALDOS DOS MATERIAIS ==========
            materiais_principais = [
                "Cimento", "Brita 0", "Brita 1", "Areia Média", 
                "Pó de Brita", "Aditivo", "Água"
            ]
            
            if self.check_todos_materiais.isChecked():
                cursor.execute("SELECT material FROM estoque WHERE material != 'Areia' ORDER BY material")
                materiais = [row[0] for row in cursor.fetchall()]
            else:
                materiais = materiais_principais
            
            # Atualizar tabela
            self.tabela_diario.setRowCount(len(materiais))
            
            consumo_total = 0
            estoque_inicial_total = 0
            estoque_final_total = 0
            
            # Determinar datas para cálculo do período
            hoje = datetime.now().strftime("%Y-%m-%d")
            data_inicio = "2000-01-01"
            
            # Ajustar data_inicio baseado nos filtros
            if filtro_ano != "Todos os anos":
                ano = filtro_ano
                if filtro_mes:
                    mes = f"{filtro_mes:02d}"
                    if filtro_dia != "Todos os dias":
                        dia = f"{int(filtro_dia):02d}"
                        data_inicio = f"{ano}-{mes}-{dia}"
                    else:
                        data_inicio = f"{ano}-{mes}-01"
                else:
                    data_inicio = f"{ano}-01-01"
            
            for i, material in enumerate(materiais):
                # Obter dados do material
                cursor.execute('''
                    SELECT quantidade, unidade, estoque_minimo 
                    FROM estoque 
                    WHERE material = ?
                ''', (material,))
                
                material_data = cursor.fetchone()
                if not material_data:
                    continue
                    
                estoque_atual = material_data[0] or 0
                unidade = material_data[1] or "kg"
                estoque_minimo = material_data[2] or 0
                
                # Calcular consumo do período
                consumo = self.obter_consumo_material_periodo(material, data_inicio, hoje)
                
                # Calcular estoque inicial (estoque atual + consumo do período)
                estoque_inicial = self.obter_estoque_inicial_periodo(material, data_inicio)
                estoque_final = estoque_atual
                
                # Calcular variação percentual
                variacao = 0
                if estoque_inicial > 0:
                    variacao = ((estoque_final - estoque_inicial) / estoque_inicial) * 100
                
                # Adicionar à tabela
                self.tabela_diario.setItem(i, 0, QTableWidgetItem(material))
                self.tabela_diario.setItem(i, 1, QTableWidgetItem(unidade))
                self.tabela_diario.setItem(i, 2, QTableWidgetItem(f"{estoque_inicial:,.2f}"))
                self.tabela_diario.setItem(i, 3, QTableWidgetItem(f"{consumo:,.2f}"))
                self.tabela_diario.setItem(i, 4, QTableWidgetItem(f"{estoque_final:,.2f}"))
                
                # Item de variação com cor
                variacao_item = QTableWidgetItem(f"{variacao:+.2f}%")
                if variacao < 0:
                    variacao_item.setForeground(QColor("#F44336"))
                elif variacao > 0:
                    variacao_item.setForeground(QColor("#4CAF50"))
                self.tabela_diario.setItem(i, 5, variacao_item)
                
                # Atualizar totais
                consumo_total += consumo
                estoque_inicial_total += estoque_inicial
                estoque_final_total += estoque_final
                
                # Atualizar labels dos materiais
                if material in materiais_principais and material in self.materiais_labels:
                    label = self.materiais_labels[material]
                    label.setText(f"{estoque_final:,.2f}")
                    
                    if estoque_final <= 0:
                        label.setStyleSheet("color: #F44336; font-weight: bold;")
                    elif estoque_final < estoque_minimo:
                        label.setStyleSheet("color: #FF9800; font-weight: bold;")
                    else:
                        label.setStyleSheet(f"color: {self.get_cor_material(material)}; font-weight: bold;")
            
            conn.close()
            
            # ========== 3. RESUMO ==========
            self.label_resumo_diario.setText(
                f"📊 {periodo_str}\n"
                f"• Volume produzido: {volume_total:,.2f} m³\n"
                f"• Consumo total: {consumo_total:,.2f} unidades\n"
                f"• Variação total do estoque: {estoque_final_total - estoque_inicial_total:+,.2f} unidades"
            )
            
        except Exception as e:
            error_msg = f"Erro ao gerar relatório diário: {str(e)}"
            print(error_msg)
            QMessageBox.warning(self, "Erro", error_msg)
    
    def gerar_relatorio_semanal(self):
        """Gera relatório semanal COM FILTROS DIA, MÊS, ANO"""
        try:
            # Obter filtros
            filtro_dia = self.combo_semanal_dia.currentText()
            filtro_mes = self.combo_semanal_mes.currentData()
            filtro_ano = self.combo_semanal_ano.currentText()
            
            # Construir período de filtro
            periodo_str = "Período: "
            conditions = []
            params = []
            
            if filtro_dia != "Todos os dias":
                conditions.append("CAST(strftime('%d', data_pesagem) AS INTEGER) = ?")
                params.append(int(filtro_dia))
                periodo_str += f"Dia {filtro_dia}, "
            
            if filtro_mes:
                conditions.append("CAST(strftime('%m', data_pesagem) AS INTEGER) = ?")
                params.append(filtro_mes)
                meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
                periodo_str += f"{meses[filtro_mes-1]}, "
            
            if filtro_ano != "Todos os anos":
                conditions.append("strftime('%Y', data_pesagem) = ?")
                params.append(filtro_ano)
                periodo_str += f"Ano {filtro_ano}"
            
            # Remover vírgula extra no final
            if periodo_str.endswith(", "):
                periodo_str = periodo_str[:-2]
            
            if periodo_str == "Período: ":
                periodo_str = "Período: Todos os dias"
            
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Query para obter dados semanalmente agrupados por dia
            query = '''
                SELECT 
                    date(p.data_pesagem) as data,
                    SUM(p.quantidade) as volume_total,
                    COUNT(DISTINCT p.cliente_id) as clientes,
                    SUM(
                        CASE WHEN json_extract(p.materiais_json, '$.Cimento.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Cimento.quantidade') 
                        ELSE 0 END
                    ) as cimento,
                    SUM(
                        CASE WHEN json_extract(p.materiais_json, '$.Brita 0.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Brita 0.quantidade') 
                        ELSE 0 END +
                        CASE WHEN json_extract(p.materiais_json, '$.Brita 1.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Brita 1.quantidade') 
                        ELSE 0 END
                    ) as brita_total,
                    SUM(
                        CASE WHEN json_extract(p.materiais_json, '$.Areia Média.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Areia Média.quantidade') 
                        ELSE 0 END
                    ) as areia_total,
                    SUM(
                        CASE WHEN json_extract(p.materiais_json, '$.Água.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Água.quantidade') / 1000.0
                        ELSE 0 END
                    ) as agua_total
                FROM pesagens p
                WHERE p.status = 'CONCLUÍDO'
            '''
            
            if conditions:
                query += " AND " + " AND ".join(conditions)
            
            query += " GROUP BY date(p.data_pesagem) ORDER BY data DESC LIMIT 7"
            
            cursor.execute(query, params)
            dados = cursor.fetchall()
            conn.close()
            
            # Atualizar tabela
            self.tabela_semanal.setRowCount(len(dados))
            
            total_volume = 0
            total_clientes = 0
            total_cimento = 0
            total_brita = 0
            total_areia = 0
            total_faturamento = 0
            
            for i, (data, volume, clientes, cimento, brita_total, areia_total, agua_total) in enumerate(dados):
                # Calcular consumo total e faturamento
                consumo_total = (cimento or 0) + (brita_total or 0) + (areia_total or 0) + (agua_total or 0)
                faturamento = volume * 250
                
                # Atualizar totais
                total_volume += volume
                total_clientes += clientes
                total_cimento += cimento or 0
                total_brita += brita_total or 0
                total_areia += areia_total or 0
                total_faturamento += faturamento
                
                # Adicionar à tabela
                self.tabela_semanal.setItem(i, 0, QTableWidgetItem(str(data)))
                self.tabela_semanal.setItem(i, 1, QTableWidgetItem(f"{volume:,.2f}"))
                self.tabela_semanal.setItem(i, 2, QTableWidgetItem(str(clientes)))
                self.tabela_semanal.setItem(i, 3, QTableWidgetItem(f"{consumo_total:,.0f}"))
                self.tabela_semanal.setItem(i, 4, QTableWidgetItem(f"{cimento:,.0f}" if cimento else "0"))
                self.tabela_semanal.setItem(i, 5, QTableWidgetItem(f"{brita_total:,.0f}" if brita_total else "0"))
                self.tabela_semanal.setItem(i, 6, QTableWidgetItem(f"{areia_total:,.0f}" if areia_total else "0"))
                self.tabela_semanal.setItem(i, 7, QTableWidgetItem(f"R$ {faturamento:,.2f}"))
            
            # Atualizar resumo
            self.label_semana_info.setText(f"📅 Semana: {periodo_str}")
            self.label_total_semanal.setText(f"📦 Total da Semana: {total_volume:,.2f} m³")
            self.label_clientes_semanal.setText(f"👥 Clientes: {total_clientes}")
            self.label_faturamento_semanal.setText(f"💰 Faturamento: R$ {total_faturamento:,.2f}")
            
        except Exception as e:
            error_msg = f"Erro ao gerar relatório semanal: {str(e)}"
            print(error_msg)
            QMessageBox.warning(self, "Erro", error_msg)
    
    def gerar_relatorio_mensal(self):
        """Gera relatório mensal COM FILTROS COMPLETOS"""
        try:
            # Obter filtros
            data_especifica = self.date_especifica_mensal.date().toString("yyyy-MM-dd")
            filtro_dia = self.combo_mensal_dia.currentText()
            filtro_semana = self.combo_mensal_semana.currentText()
            filtro_mes = self.combo_mensal_mes.currentData()
            filtro_ano = self.combo_mensal_ano.currentText()
            
            # Construir período de filtro
            periodo_str = "Período: "
            conditions = []
            params = []
            
            # Data Específica
            if self.date_especifica_mensal.date() != QDate.currentDate():
                conditions.append("date(data_pesagem) = ?")
                params.append(data_especifica)
                periodo_str += f"Data específica: {self.date_especifica_mensal.date().toString('dd/MM/yyyy')}, "
            
            # Dia do Mês
            if filtro_dia != "Todos os dias":
                conditions.append("CAST(strftime('%d', data_pesagem) AS INTEGER) = ?")
                params.append(int(filtro_dia))
                periodo_str += f"Dia {filtro_dia}, "
            
            # Número da Semana
            if filtro_semana != "Todas as semanas":
                conditions.append("CAST(strftime('%W', data_pesagem) AS INTEGER) = ?")
                params.append(int(filtro_semana))
                periodo_str += f"Semana {filtro_semana}, "
            
            # Mês
            if filtro_mes:
                conditions.append("CAST(strftime('%m', data_pesagem) AS INTEGER) = ?")
                params.append(filtro_mes)
                meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
                periodo_str += f"{meses[filtro_mes-1]}, "
            
            # Ano
            if filtro_ano != "Todos os anos":
                conditions.append("strftime('%Y', data_pesagem) = ?")
                params.append(filtro_ano)
                periodo_str += f"Ano {filtro_ano}"
            
            # Remover vírgula extra no final
            if periodo_str.endswith(", "):
                periodo_str = periodo_str[:-2]
            
            if periodo_str == "Período: ":
                periodo_str = "Período: Todos os registros"
            
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Query para obter dados mensais agrupados
            query = '''
                SELECT 
                    date(p.data_pesagem) as data,
                    CAST(strftime('%d', p.data_pesagem) AS INTEGER) as dia,
                    CAST(strftime('%W', p.data_pesagem) AS INTEGER) as semana,
                    strftime('%m', p.data_pesagem) as mes_num,
                    strftime('%Y', p.data_pesagem) as ano,
                    SUM(p.quantidade) as volume_total,
                    COUNT(DISTINCT p.cliente_id) as clientes,
                    SUM(
                        CASE WHEN json_extract(p.materiais_json, '$.Cimento.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Cimento.quantidade') 
                        ELSE 0 END +
                        CASE WHEN json_extract(p.materiais_json, '$.Brita 0.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Brita 0.quantidade') 
                        ELSE 0 END +
                        CASE WHEN json_extract(p.materiais_json, '$.Brita 1.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Brita 1.quantidade') 
                        ELSE 0 END +
                        CASE WHEN json_extract(p.materiais_json, '$.Areia Média.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Areia Média.quantidade') 
                        ELSE 0 END +
                        CASE WHEN json_extract(p.materiais_json, '$.Água.quantidade') IS NOT NULL 
                        THEN json_extract(p.materiais_json, '$.Água.quantidade') / 1000.0
                        ELSE 0 END
                    ) as consumo_total,
                    SUM(p.quantidade * 150) as custo_total,
                    SUM(p.quantidade * 250) as faturamento
                FROM pesagens p
                WHERE p.status = 'CONCLUÍDO'
            '''
            
            if conditions:
                query += " AND " + " AND ".join(conditions)
            
            query += " GROUP BY date(p.data_pesagem) ORDER BY data DESC LIMIT 30"
            
            cursor.execute(query, params)
            dados = cursor.fetchall()
            conn.close()
            
            # Atualizar tabela
            self.tabela_mensal.setRowCount(len(dados))
            
            total_volume = 0
            total_clientes = 0
            total_consumo = 0
            total_custo = 0
            total_faturamento = 0
            
            meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                          "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
            
            for i, (data, dia, semana, mes_num, ano, volume, clientes, consumo, custo, faturamento) in enumerate(dados):
                # Atualizar totais
                total_volume += volume or 0
                total_clientes += clientes or 0
                total_consumo += consumo or 0
                total_custo += custo or 0
                total_faturamento += faturamento or 0
                
                # Nome do mês
                mes_nome = meses_nomes[int(mes_num) - 1] if mes_num else ""
                
                # Adicionar à tabela
                self.tabela_mensal.setItem(i, 0, QTableWidgetItem(str(data)))
                self.tabela_mensal.setItem(i, 1, QTableWidgetItem(str(dia)))
                self.tabela_mensal.setItem(i, 2, QTableWidgetItem(str(semana)))
                self.tabela_mensal.setItem(i, 3, QTableWidgetItem(mes_nome))
                self.tabela_mensal.setItem(i, 4, QTableWidgetItem(str(ano)))
                self.tabela_mensal.setItem(i, 5, QTableWidgetItem(f"{volume:,.2f}" if volume else "0.00"))
                self.tabela_mensal.setItem(i, 6, QTableWidgetItem(str(clientes) if clientes else "0"))
                self.tabela_mensal.setItem(i, 7, QTableWidgetItem(f"{consumo:,.0f}" if consumo else "0"))
                self.tabela_mensal.setItem(i, 8, QTableWidgetItem(f"R$ {custo:,.2f}" if custo else "R$ 0,00"))
                self.tabela_mensal.setItem(i, 9, QTableWidgetItem(f"R$ {faturamento:,.2f}" if faturamento else "R$ 0,00"))
            
            # Atualizar resumo
            lucro = total_faturamento - total_custo
            lucro_percentual = (lucro / total_custo * 100) if total_custo > 0 else 0
            
            self.label_periodo_info.setText(f"📅 {periodo_str}")
            self.label_total_mensal.setText(f"📦 Volume Total: {total_volume:,.2f} m³")
            self.label_clientes_mensal.setText(f"👥 Clientes: {total_clientes}")
            self.label_faturamento_mensal.setText(f"💰 Faturamento: R$ {total_faturamento:,.2f} | Lucro: R$ {lucro:,.2f} ({lucro_percentual:.1f}%)")
            
        except Exception as e:
            error_msg = f"Erro ao gerar relatório mensal: {str(e)}"
            print(error_msg)
            QMessageBox.warning(self, "Erro", error_msg)
    
    def gerar_relatorio_gastos_corrigido(self):
        """Gera relatório de gastos COM FILTROS COMPLETOS - VERSÃO CORRIGIDA"""
        try:
            # Obter filtros
            data_especifica = self.date_especifica_gastos.date().toString("yyyy-MM-dd")
            filtro_dia = self.combo_gastos_dia.currentText()
            filtro_semana = self.combo_gastos_semana.currentText()
            filtro_mes = self.combo_gastos_mes.currentData()
            filtro_ano = self.combo_gastos_ano.currentText()
            
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            # Primeiro, verificar a estrutura da tabela
            cursor.execute("PRAGMA table_info(historico_estoque)")
            colunas = cursor.fetchall()
            colunas_nomes = [col[1] for col in colunas]
            
            # Construir a query dinamicamente baseado nas colunas disponíveis
            campos_select = [
                "date(data_movimentacao) as data",
                "CAST(strftime('%d', data_movimentacao) AS INTEGER) as dia",
                "CAST(strftime('%W', data_movimentacao) AS INTEGER) as semana",
                "strftime('%m', data_movimentacao) as mes_num",
                "strftime('%Y', data_movimentacao) as ano",
                "destino as fornecedor",
                "material",
                "quantidade",
                "unidade"
            ]
            
            # Verificar se a coluna valor_unitario existe
            if "valor_unitario" in colunas_nomes:
                campos_select.append("valor_unitario")
                campos_select.append("(quantidade * valor_unitario) as valor_total")
            else:
                # Se não existir, usar valor padrão
                campos_select.append("0 as valor_unitario")
                campos_select.append("0 as valor_total")
            
            # Construir query base
            query = f'''
                SELECT {', '.join(campos_select)}
                FROM historico_estoque 
                WHERE tipo = 'ENTRADA'
            '''
            
            params = []
            conditions = []
            
            # Aplicar filtros
            # Data Específica
            if self.date_especifica_gastos.date() != QDate.currentDate():
                conditions.append("date(data_movimentacao) = ?")
                params.append(data_especifica)
            
            # Dia do Mês
            if filtro_dia != "Todos os dias":
                conditions.append("CAST(strftime('%d', data_movimentacao) AS INTEGER) = ?")
                params.append(int(filtro_dia))
            
            # Número da Semana
            if filtro_semana != "Todas as semanas":
                conditions.append("CAST(strftime('%W', data_movimentacao) AS INTEGER) = ?")
                params.append(int(filtro_semana))
            
            # Mês
            if filtro_mes:
                conditions.append("CAST(strftime('%m', data_movimentacao) AS INTEGER) = ?")
                params.append(filtro_mes)
            
            # Ano
            if filtro_ano != "Todos os anos":
                conditions.append("strftime('%Y', data_movimentacao) = ?")
                params.append(filtro_ano)
            
            if conditions:
                query += " AND " + " AND ".join(conditions)
            
            query += " ORDER BY data_movimentacao DESC"
            
            cursor.execute(query, params)
            dados = cursor.fetchall()
            conn.close()
            
            # Atualizar tabela
            self.tabela_gastos.setRowCount(len(dados))
            
            total_gastos = 0
            meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
                          "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
            
            for i, row in enumerate(dados):
                # Extrair dados da linha - a estrutura depende das colunas disponíveis
                data = row[0]
                dia = row[1]
                semana = row[2]
                mes_num = row[3]
                ano = row[4]
                fornecedor = row[5]
                material = row[6]
                quantidade = row[7]
                unidade = row[8]
                valor_unitario = row[9] if len(row) > 9 else 0
                valor_total = row[10] if len(row) > 10 else 0
                
                # Atualizar total
                total_gastos += valor_total or 0
                
                # Nome do mês
                mes_nome = meses_nomes[int(mes_num) - 1] if mes_num else ""
                
                # Adicionar à tabela
                self.tabela_gastos.setItem(i, 0, QTableWidgetItem(str(data)))
                self.tabela_gastos.setItem(i, 1, QTableWidgetItem(str(dia)))
                self.tabela_gastos.setItem(i, 2, QTableWidgetItem(str(semana)))
                self.tabela_gastos.setItem(i, 3, QTableWidgetItem(mes_nome))
                self.tabela_gastos.setItem(i, 4, QTableWidgetItem(str(ano)))
                self.tabela_gastos.setItem(i, 5, QTableWidgetItem(fornecedor or "N/A"))
                self.tabela_gastos.setItem(i, 6, QTableWidgetItem(material or "N/A"))
                self.tabela_gastos.setItem(i, 7, QTableWidgetItem(f"{quantidade:,.2f}" if quantidade else "0.00"))
                self.tabela_gastos.setItem(i, 8, QTableWidgetItem(unidade or ""))
                self.tabela_gastos.setItem(i, 9, QTableWidgetItem(f"R$ {valor_unitario:,.2f}" if valor_unitario else "R$ 0,00"))
                self.tabela_gastos.setItem(i, 10, QTableWidgetItem(f"R$ {valor_total:,.2f}" if valor_total else "R$ 0,00"))
            
            # Atualizar total de gastos
            self.label_total_gastos.setText(f"💰 Total de Gastos: R$ {total_gastos:,.2f}")
            
        except sqlite3.OperationalError as e:
            error_msg = str(e)
            QMessageBox.warning(self, "Erro no Banco de Dados", 
                f"Erro ao acessar banco de dados: {error_msg}\n\n"
                "Solução: Execute o diagnóstico (botão '🔧 Diagnóstico') para verificar "
                "a estrutura do banco e criar a coluna 'valor_unitario' se necessário.")
            
            # Mostrar dados mesmo sem valor_unitario
            self.label_total_gastos.setText("💰 Total de Gastos: R$ 0,00 (coluna 'valor_unitario' não encontrada)")
            
        except Exception as e:
            error_msg = f"Erro ao gerar relatório de gastos: {str(e)}"
            print(error_msg)
            QMessageBox.warning(self, "Erro", error_msg)
    
    # ========== FUNÇÕES DE EXPORTAÇÃO E IMPRESSÃO ==========
    
    def get_dados_tabela_atual(self):
        """Obtém os dados da tabela atual baseado na aba selecionada"""
        tab_index = self.tabs.currentIndex()
        nome_aba = self.tabs.tabText(tab_index)
        
        dados = {
            'titulo': f"RELATÓRIO - {nome_aba} - USINA BETTO MIX",
            'cabecalhos': [],
            'linhas': [],
            'info_adicional': []
        }
        
        try:
            if tab_index == 0:  # DIÁRIO
                tabela = self.tabela_diario
                dados['cabecalhos'] = ['Material', 'Unidade', 'Estoque Inicial', 'Consumo', 'Estoque Final', 'Variação %']
                
                for linha in range(tabela.rowCount()):
                    linha_dados = []
                    for coluna in range(tabela.columnCount()):
                        item = tabela.item(linha, coluna)
                        linha_dados.append(item.text() if item else "")
                    dados['linhas'].append(linha_dados)
                
                # Adicionar informações adicionais
                dados['info_adicional'].append(f"Volume total: {self.label_volume_total.text()}")
                dados['info_adicional'].append(f"Número de pesagens: {self.label_num_pesagens.text()}")
                dados['info_adicional'].append(f"Traços utilizados: {self.label_tracos_utilizados.text()}")
                dados['info_adicional'].append(f"Clientes atendidos: {self.label_clientes_atendidos.text()}")
                dados['info_adicional'].append(f"Resumo: {self.label_resumo_diario.text()}")
                
            elif tab_index == 1:  # PRODUÇÃO - COM FILTROS AVANÇADOS
                tabela = self.tabela_producao_filtrada
                dados['cabecalhos'] = ['Data', 'Volume (m³)', 'Pesagem', 'Traço', 'Cliente', 
                                       'M³ Médio', 'Status', 'Hora', 'Motorista', 'Placa', 'Observação']
                
                for linha in range(tabela.rowCount()):
                    linha_dados = []
                    for coluna in range(tabela.columnCount()):
                        item = tabela.item(linha, coluna)
                        linha_dados.append(item.text() if item else "")
                    dados['linhas'].append(linha_dados)
                
                # Adicionar informações adicionais
                dados['info_adicional'].append(f"Volume Total: {self.label_total_volume_filtrado.text()}")
                dados['info_adicional'].append(f"Total de Pesagens: {self.label_total_pesagens_filtrado.text()}")
                dados['info_adicional'].append(f"M³ Médio: {self.label_media_m3_filtrado.text()}")
                dados['info_adicional'].append(f"Status: {self.label_status_geral_filtrado.text()}")
                dados['info_adicional'].append(f"Período: {self.date_inicial_producao.date().toString('dd/MM/yyyy')} a {self.date_final_producao.date().toString('dd/MM/yyyy')}")
                
            elif tab_index == 2:  # SEMANAL
                tabela = self.tabela_semanal
                dados['cabecalhos'] = ['Dia', 'Produção (m³)', 'Clientes', 'Consumo Total', 'Cimento (kg)', 'Brita (kg)', 'Areia (kg)', 'Faturamento']
                
                for linha in range(tabela.rowCount()):
                    linha_dados = []
                    for coluna in range(tabela.columnCount()):
                        item = tabela.item(linha, coluna)
                        linha_dados.append(item.text() if item else "")
                    dados['linhas'].append(linha_dados)
                
                # Adicionar informações adicionais
                dados['info_adicional'].append(f"Informações da Semana: {self.label_semana_info.text()}")
                dados['info_adicional'].append(f"Total da Semana: {self.label_total_semanal.text()}")
                dados['info_adicional'].append(f"Clientes: {self.label_clientes_semanal.text()}")
                dados['info_adicional'].append(f"Faturamento: {self.label_faturamento_semanal.text()}")
                
            elif tab_index == 3:  # MENSAL
                tabela = self.tabela_mensal
                dados['cabecalhos'] = ['Data', 'Dia', 'Semana', 'Mês', 'Ano', 'Produção (m³)', 'Clientes', 'Consumo Total (kg)', 'Custo Total (R$)', 'Faturamento (R$)']
                
                for linha in range(tabela.rowCount()):
                    linha_dados = []
                    for coluna in range(tabela.columnCount()):
                        item = tabela.item(linha, coluna)
                        linha_dados.append(item.text() if item else "")
                    dados['linhas'].append(linha_dados)
                
                # Adicionar informações adicionais
                dados['info_adicional'].append(f"Período: {self.label_periodo_info.text()}")
                dados['info_adicional'].append(f"Volume Total: {self.label_total_mensal.text()}")
                dados['info_adicional'].append(f"Clientes: {self.label_clientes_mensal.text()}")
                dados['info_adicional'].append(f"Faturamento: {self.label_faturamento_mensal.text()}")
                
            elif tab_index == 4:  # GASTOS
                tabela = self.tabela_gastos
                dados['cabecalhos'] = ['Data', 'Dia', 'Semana', 'Mês', 'Ano', 'Fornecedor', 'Material', 'Quantidade', 'Unidade', 'Valor Unitário (R$)', 'Valor Total (R$)']
                
                for linha in range(tabela.rowCount()):
                    linha_dados = []
                    for coluna in range(tabela.columnCount()):
                        item = tabela.item(linha, coluna)
                        linha_dados.append(item.text() if item else "")
                    dados['linhas'].append(linha_dados)
                
                # Adicionar informações adicionais
                dados['info_adicional'].append(f"Total de Gastos: {self.label_total_gastos.text()}")
                
            elif tab_index == 5:  # ESTOQUE
                tabela = self.tabela_estoque_detalhado
                dados['cabecalhos'] = ['Material', 'Quantidade', 'Unidade', 'Estoque Mínimo', 'Estoque Ideal', 'Status', 'Última Atualização']
                
                for linha in range(tabela.rowCount()):
                    linha_dados = []
                    for coluna in range(tabela.columnCount()):
                        item = tabela.item(linha, coluna)
                        linha_dados.append(item.text() if item else "")
                    dados['linhas'].append(linha_dados)
                
                # Adicionar informações adicionais
                dados['info_adicional'].append("Estoque Ideal fixo em 5.000 para todos os materiais")
                dados['info_adicional'].append("Material 'Areia' excluído conforme solicitado")
                dados['info_adicional'].append(f"Total de itens: {tabela.rowCount()}")
                
        except Exception as e:
            print(f"Erro ao coletar dados: {str(e)}")
        
        return dados
    
    def exportar_excel(self):
        """Exporta o relatório da aba atual para Excel"""
        try:
            # Obter dados da tabela atual
            dados = self.get_dados_tabela_atual()
            
            if not dados['linhas']:
                QMessageBox.warning(self, "Aviso", "Não há dados para exportar!")
                return
            
            # Abrir diálogo para salvar arquivo
            data_atual = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_aba = self.tabs.tabText(self.tabs.currentIndex()).replace(" ", "_").replace("📅", "DIARIO").replace("⚙️", "PRODUCAO").replace("📈", "SEMANAL").replace("📊", "MENSAL").replace("💰", "GASTOS").replace("📦", "ESTOQUE")
            nome_sugerido = f"Relatorio_{nome_aba}_{data_atual}.xlsx"
            
            filename, _ = QFileDialog.getSaveFileName(
                self, 
                "Salvar Relatório Excel", 
                nome_sugerido,
                "Arquivos Excel (*.xlsx);;Todos os arquivos (*)"
            )
            
            if not filename:
                return
            
            # Garantir extensão .xlsx
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            # Criar progress dialog
            progress = QProgressDialog("Exportando para Excel...", "Cancelar", 0, 100, self)
            progress.setWindowTitle("Exportação")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.show()
            
            # Atualizar progresso
            QCoreApplication.processEvents()
            progress.setValue(10)
            
            # Criar workbook do Excel
            wb = Workbook()
            ws = wb.active
            ws.title = f"Relatório {nome_aba}"
            
            # Estilos
            header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            title_font = Font(size=14, bold=True)
            info_font = Font(size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:K1')
            title_cell = ws['A1']
            title_cell.value = dados['titulo']
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center')
            
            # Data de exportação
            ws['A2'] = f"Exportado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            ws['A2'].font = info_font
            
            # Informações adicionais
            linha_atual = 3
            for info in dados['info_adicional']:
                if info:
                    ws.cell(row=linha_atual, column=1).value = info
                    ws.cell(row=linha_atual, column=1).font = info_font
                    linha_atual += 1
            
            linha_atual += 1
            
            # Cabeçalhos
            for col_idx, cabecalho in enumerate(dados['cabecalhos'], 1):
                cell = ws.cell(row=linha_atual, column=col_idx)
                cell.value = cabecalho
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            progress.setValue(30)
            QCoreApplication.processEvents()
            
            # Dados
            for linha_idx, linha_dados in enumerate(dados['linhas'], 1):
                for col_idx, valor in enumerate(linha_dados, 1):
                    cell = ws.cell(row=linha_atual + linha_idx, column=col_idx)
                    cell.value = valor
                    cell.border = border
                    
                    # Formatação especial para números e valores
                    if any(x in dados['cabecalhos'][col_idx-1].lower() for x in ['quantidade', 'valor', 'custo', 'faturamento', 'consumo', 'estoque', 'volume', 'm³']):
                        try:
                            valor_limpo = str(valor).replace('.', '').replace(',', '.')
                            num_valor = float(valor_limpo)
                            cell.value = num_valor
                            
                            # Formatar como número com 2 casas decimais
                            if 'R$' in dados['cabecalhos'][col_idx-1]:
                                cell.number_format = '"R$" #,##0.00'
                            else:
                                cell.number_format = '#,##0.00'
                        except:
                            pass
                
                # Atualizar progresso
                if linha_idx % 10 == 0:
                    progress.setValue(30 + (linha_idx * 40 // len(dados['linhas'])))
                    QCoreApplication.processEvents()
                    if progress.wasCanceled():
                        return
            
            progress.setValue(80)
            QCoreApplication.processEvents()
            
            # Ajustar largura das colunas
            for col_idx in range(1, len(dados['cabecalhos']) + 1):
                max_length = 0
                column = get_column_letter(col_idx)
                
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Adicionar fórmulas de totais se aplicável
            ultima_linha = linha_atual + len(dados['linhas'])
            if 'Valor Total (R$)' in dados['cabecalhos']:
                col_idx = dados['cabecalhos'].index('Valor Total (R$)') + 1
                col_letter = get_column_letter(col_idx)
                ws.cell(row=ultima_linha, column=col_idx).value = f"=SUM({col_letter}{linha_atual+1}:{col_letter}{ultima_linha-1})"
                ws.cell(row=ultima_linha, column=col_idx).number_format = '"R$" #,##0.00'
                ws.cell(row=ultima_linha, column=col_idx).font = Font(bold=True)
                ws.cell(row=ultima_linha, column=1).value = "TOTAL GERAL"
                ws.cell(row=ultima_linha, column=1).font = Font(bold=True)
            
            # Congelar painel (cabeçalhos)
            ws.freeze_panes = ws[f'A{linha_atual+1}']
            
            progress.setValue(90)
            QCoreApplication.processEvents()
            
            # Salvar arquivo
            wb.save(filename)
            
            progress.setValue(100)
            QCoreApplication.processEvents()
            
            # Perguntar se quer abrir o arquivo
            resposta = QMessageBox.question(
                self,
                "Exportação Concluída",
                f"Relatório exportado com sucesso para:\n{filename}\n\nDeseja abrir o arquivo?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if resposta == QMessageBox.StandardButton.Yes:
                self.abrir_arquivo_excel(filename)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro na Exportação", f"Erro ao exportar para Excel: {str(e)}")
    
    def abrir_arquivo_excel(self, filename):
        """Abre o arquivo Excel no aplicativo padrão"""
        try:
            if sys.platform == "win32":
                os.startfile(filename)
            elif sys.platform == "darwin":
                subprocess.call(["open", filename])
            else:
                subprocess.call(["xdg-open", filename])
        except Exception as e:
            QMessageBox.warning(self, "Aviso", f"Não foi possível abrir o arquivo: {str(e)}")
    
    def imprimir_relatorio_corrigido(self):
        """Imprime o relatório da aba atual - VERSÃO CORRIGIDA"""
        try:
            # Obter dados da tabela atual
            dados = self.get_dados_tabela_atual()
            
            if not dados['linhas']:
                QMessageBox.warning(self, "Aviso", "Não há dados para imprimir!")
                return
            
            # CORREÇÃO: Criar QMessageBox personalizado com botões customizados
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Imprimir Relatório")
            msg_box.setText("Escolha o formato de impressão:")
            msg_box.setIcon(QMessageBox.Icon.Question)
            
            # Adicionar botões customizados usando addButton (não StandardButton)
            btn_print = msg_box.addButton("📄 Imprimir", QMessageBox.ButtonRole.ActionRole)
            btn_pdf = msg_box.addButton("💾 Salvar PDF", QMessageBox.ButtonRole.ActionRole)
            btn_cancel = msg_box.addButton(QMessageBox.StandardButton.Cancel)
            
            # Executar a caixa de diálogo
            msg_box.exec()
            
            # Verificar qual botão foi pressionado
            clicked_button = msg_box.clickedButton()
            
            if clicked_button == btn_cancel:
                return
            elif clicked_button == btn_print or clicked_button == btn_pdf:
                # Criar arquivo HTML para impressão/PDF
                html_content = self.criar_html_impressao(dados)
                
                if clicked_button == btn_print:
                    self.imprimir_html(html_content)
                else:
                    self.salvar_pdf(html_content, dados)
                    
        except Exception as e:
            QMessageBox.critical(self, "Erro na Impressão", f"Erro ao imprimir relatório: {str(e)}")
    
    def criar_html_impressao(self, dados):
        """Cria conteúdo HTML para impressão"""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>{dados['titulo']}</title>
            <style>
                body {{
                    font-family: 'Arial', sans-serif;
                    margin: 20px;
                    color: #333;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 30px;
                    border-bottom: 2px solid #0D47A1;
                    padding-bottom: 20px;
                }}
                .title {{
                    font-size: 24px;
                    font-weight: bold;
                    color: #0D47A1;
                    margin-bottom: 10px;
                }}
                .subtitle {{
                    font-size: 16px;
                    color: #666;
                    margin-bottom: 20px;
                }}
                .info {{
                    background-color: #f5f5f5;
                    padding: 15px;
                    margin-bottom: 20px;
                    border-radius: 5px;
                    border-left: 4px solid #0D47A1;
                }}
                .info-item {{
                    margin-bottom: 5px;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin-bottom: 30px;
                }}
                th {{
                    background-color: #0D47A1;
                    color: white;
                    font-weight: bold;
                    padding: 12px;
                    text-align: left;
                    border: 1px solid #0D47A1;
                }}
                td {{
                    padding: 10px;
                    border: 1px solid #ddd;
                }}
                tr:nth-child(even) {{
                    background-color: #f9f9f9;
                }}
                tr:hover {{
                    background-color: #f5f5f5;
                }}
                .total-row {{
                    font-weight: bold;
                    background-color: #e3f2fd;
                }}
                .footer {{
                    margin-top: 30px;
                    text-align: center;
                    font-size: 12px;
                    color: #666;
                    border-top: 1px solid #ddd;
                    padding-top: 20px;
                }}
                .print-date {{
                    margin-bottom: 20px;
                    text-align: right;
                    font-style: italic;
                }}
                @media print {{
                    body {{
                        margin: 0;
                        padding: 0;
                    }}
                    .no-print {{
                        display: none;
                    }}
                    table {{
                        page-break-inside: auto;
                    }}
                    tr {{
                        page-break-inside: avoid;
                        page-break-after: auto;
                    }}
                    thead {{
                        display: table-header-group;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <div class="title">{dados['titulo']}</div>
                <div class="subtitle">USINA BETTO MIX - SISTEMA DE GESTÃO</div>
                <div class="print-date">Impresso em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</div>
            </div>
        """
        
        # Adicionar informações adicionais
        if dados['info_adicional']:
            html += '<div class="info">'
            for info in dados['info_adicional']:
                if info:
                    html += f'<div class="info-item">• {info}</div>'
            html += '</div>'
        
        # Adicionar tabela
        html += '<table>'
        
        # Cabeçalhos
        html += '<thead><tr>'
        for cabecalho in dados['cabecalhos']:
            html += f'<th>{cabecalho}</th>'
        html += '</tr></thead>'
        
        # Dados
        html += '<tbody>'
        for linha in dados['linhas']:
            html += '<tr>'
            for valor in linha:
                # Verificar se é linha de total
                if linha[0] in ['TOTAL', 'TOTAL GERAL', 'TOTAL DA SEMANA']:
                    html += f'<td class="total-row">{valor}</td>'
                else:
                    html += f'<td>{valor}</td>'
            html += '</tr>'
        html += '</tbody>'
        html += '</table>'
        
        # Rodapé
        html += f"""
            <div class="footer">
                <p>USINA BETTO MIX | Sistema de Gestão de Concreto</p>
                <p>Documento gerado automaticamente pelo sistema</p>
                <p>Página 1 de 1</p>
            </div>
            
            <div class="no-print" style="text-align: center; margin-top: 30px;">
                <button onclick="window.print()" style="
                    background-color: #0D47A1;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 5px;
                    cursor: pointer;
                    font-size: 14px;
                ">🖨️ Imprimir Documento</button>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def imprimir_html(self, html_content):
        """Imprime o conteúdo HTML"""
        try:
            # Criar arquivo temporário HTML
            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                f.write(html_content)
                temp_file = f.name
            
            # Abrir no navegador para impressão
            self.abrir_arquivo_html(temp_file)
            
        except Exception as e:
            QMessageBox.warning(self, "Erro", f"Erro ao preparar impressão: {str(e)}")
    
    def salvar_pdf(self, html_content, dados):
        """Salva o relatório como PDF"""
        try:
            # Sugerir nome do arquivo
            data_atual = datetime.now().strftime("%Y%m%d_%H%M%S")
            nome_aba = self.tabs.tabText(self.tabs.currentIndex()).replace(" ", "_").replace("📅", "DIARIO").replace("⚙️", "PRODUCAO").replace("📈", "SEMANAL").replace("📊", "MENSAL").replace("💰", "GASTOS").replace("📦", "ESTOQUE")
            nome_sugerido = f"Relatorio_{nome_aba}_{data_atual}.html"
            
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Salvar Relatório como HTML",
                nome_sugerido,
                "Arquivos HTML (*.html);;Todos os arquivos (*)"
            )
            
            if not filename:
                return
            
            # Garantir extensão .html
            if not filename.endswith('.html'):
                filename += '.html'
            
            # Salvar arquivo HTML
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # Perguntar se quer abrir
            resposta = QMessageBox.question(
                self,
                "PDF Gerado",
                f"Relatório salvo como HTML em:\n{filename}\n\nDeseja abrir o arquivo?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            if resposta == QMessageBox.StandardButton.Yes:
                self.abrir_arquivo_html(filename)
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar PDF: {str(e)}")
    
    def abrir_arquivo_html(self, filename):
        """Abre o arquivo HTML no navegador padrão"""
        try:
            if sys.platform == "win32":
                os.startfile(filename)
            elif sys.platform == "darwin":
                subprocess.call(["open", filename])
            else:
                subprocess.call(["xdg-open", filename])
        except Exception as e:
            QMessageBox.warning(self, "Aviso", f"Não foi possível abrir o arquivo: {str(e)}\n\nO arquivo foi salvo em: {filename}")