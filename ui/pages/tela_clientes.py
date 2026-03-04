from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
    QPushButton, QTableWidget, QTableWidgetItem,
    QLineEdit, QFormLayout, QGroupBox, QMessageBox,
    QHeaderView, QDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont, QColor
import sqlite3
import re
import os
import subprocess  # Para abrir Excel automaticamente no Windows

class TelaClientes(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.cliente_selecionado = None
        self.init_ui()
        self.carregar_clientes()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Título
        titulo = QLabel("GERENCIAMENTO DE CLIENTES")
        titulo.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("color: #0D47A1; padding: 10px;")
        layout.addWidget(titulo)
        
        # Formulário
        form_group = QGroupBox("Cadastro/Edição de Cliente")
        form_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #1976D2;
                border-radius: 5px;
                margin-top: 10px;
                background-color: #E3F2FD;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                color: #0D47A1;
            }
        """)
        form_layout = QFormLayout()
        
        self.input_nome = QLineEdit()
        self.input_nome.setPlaceholderText("Nome completo ou razão social")
        self.input_nome.setStyleSheet("""
            QLineEdit {
                background-color: #FFFFFF;
                border: 1px solid #1976D2;
                border-radius: 3px;
                padding: 5px;
            }
        """)
        
        self.input_cnpj_cpf = QLineEdit()
        self.input_cnpj_cpf.setPlaceholderText("00.000.000/0000-00 ou 000.000.000-00")
        self.input_cnpj_cpf.setStyleSheet("""
            QLineEdit {
                background-color: #FFFFFF;
                border: 1px solid #1976D2;
                border-radius: 3px;
                padding: 5px;
            }
        """)
        
        self.input_endereco = QLineEdit()
        self.input_endereco.setPlaceholderText("Endereço completo")
        self.input_endereco.setStyleSheet("""
            QLineEdit {
                background-color: #FFFFFF;
                border: 1px solid #1976D2;
                border-radius: 3px;
                padding: 5px;
            }
        """)
        
        self.input_telefone = QLineEdit()
        self.input_telefone.setPlaceholderText("(11) 99999-9999")
        self.input_telefone.setStyleSheet("""
            QLineEdit {
                background-color: #FFFFFF;
                border: 1px solid #1976D2;
                border-radius: 3px;
                padding: 5px;
            }
        """)
        
        self.input_email = QLineEdit()
        self.input_email.setPlaceholderText("email@exemplo.com")
        self.input_email.setStyleSheet("""
            QLineEdit {
                background-color: #FFFFFF;
                border: 1px solid #1976D2;
                border-radius: 3px;
                padding: 5px;
            }
        """)
        
        form_layout.addRow("Nome *:", self.input_nome)
        form_layout.addRow("CNPJ/CPF:", self.input_cnpj_cpf)
        form_layout.addRow("Endereço:", self.input_endereco)
        form_layout.addRow("Telefone:", self.input_telefone)
        form_layout.addRow("Email:", self.input_email)
        
        form_group.setLayout(form_layout)
        layout.addWidget(form_group)
        
        # Botões do formulário
        btn_layout = QHBoxLayout()
        
        btn_novo = QPushButton("➕ Novo")
        btn_novo.clicked.connect(self.novo_cliente)
        btn_novo.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2196F3;
                border: 2px solid #1976D2;
            }
        """)
        
        btn_salvar = QPushButton("💾 Salvar")
        btn_salvar.clicked.connect(self.salvar_cliente)
        btn_salvar.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2196F3;
                border: 2px solid #1976D2;
            }
        """)
        
        btn_excluir = QPushButton("🗑️ Excluir")
        btn_excluir.clicked.connect(self.excluir_cliente)
        btn_excluir.setStyleSheet("""
            QPushButton {
                background-color: #1565C0;
                color: white;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
                border: 2px solid #1565C0;
            }
        """)
        
        btn_limpar = QPushButton("🧹 Limpar")
        btn_limpar.clicked.connect(self.limpar_formulario)
        btn_limpar.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2196F3;
                border: 2px solid #1976D2;
            }
        """)
        
        btn_layout.addWidget(btn_novo)
        btn_layout.addWidget(btn_salvar)
        btn_layout.addWidget(btn_excluir)
        btn_layout.addWidget(btn_limpar)
        
        layout.addLayout(btn_layout)
        
        # Tabela de clientes
        self.tabela_clientes = QTableWidget()
        self.tabela_clientes.setColumnCount(6)
        self.tabela_clientes.setHorizontalHeaderLabels([
            "ID", "Nome", "CNPJ/CPF", "Endereço", "Telefone", "Email"
        ])
        
        # Estilo da tabela
        self.tabela_clientes.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;
                alternate-background-color: #E3F2FD;
                gridline-color: #BBDEFB;
                border: 1px solid #1976D2;
                border-radius: 3px;
            }
            QTableWidget::item {
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #2196F3;
                color: white;
            }
            QHeaderView::section {
                background-color: #0D47A1;
                color: white;
                font-weight: bold;
                padding: 5px;
                border: 1px solid #1976D2;
            }
        """)
        
        self.tabela_clientes.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabela_clientes.setAlternatingRowColors(True)
        self.tabela_clientes.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabela_clientes.clicked.connect(self.selecionar_cliente)
        
        layout.addWidget(self.tabela_clientes)
        
        # Botões de ação da tabela
        btn_tabela_layout = QHBoxLayout()
        
        btn_atualizar = QPushButton("🔄 Atualizar")
        btn_atualizar.clicked.connect(self.carregar_clientes)
        btn_atualizar.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2196F3;
                border: 2px solid #1976D2;
            }
        """)
        
        btn_exportar = QPushButton("📤 Exportar")
        btn_exportar.clicked.connect(self.exportar_clientes)
        btn_exportar.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2196F3;
                border: 2px solid #1976D2;
            }
        """)
        
        btn_exportar_excel = QPushButton("📊 Exportar Excel")
        btn_exportar_excel.clicked.connect(self.exportar_clientes_excel)
        btn_exportar_excel.setStyleSheet("""
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2196F3;
                border: 2px solid #1976D2;
            }
        """)
        
        btn_imprimir = QPushButton("🖨️ Imprimir")
        btn_imprimir.clicked.connect(self.imprimir_clientes)
        btn_imprimir.setStyleSheet("""
            QPushButton {
                background-color: #1565C0;
                color: white;
                border: 2px solid #0D47A1;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
                border: 2px solid #1565C0;
            }
        """)
        
        btn_tabela_layout.addWidget(btn_atualizar)
        btn_tabela_layout.addWidget(btn_exportar)
        btn_tabela_layout.addWidget(btn_exportar_excel)
        btn_tabela_layout.addWidget(btn_imprimir)
        btn_tabela_layout.addStretch()
        
        layout.addLayout(btn_tabela_layout)
    
    # ------------------ AÇÕES ------------------ #
    
    def carregar_clientes(self):
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM clientes ORDER BY nome")
            clientes = cursor.fetchall()
            conn.close()
            
            self.tabela_clientes.setRowCount(len(clientes))
            
            for i, cliente in enumerate(clientes):
                for j in range(6):
                    item = QTableWidgetItem(str(cliente[j]) if cliente[j] else "")
                    if self.cliente_selecionado and cliente[0] == self.cliente_selecionado:
                        item.setBackground(QColor(187, 222, 251))  # Azul claro
                    self.tabela_clientes.setItem(i, j, item)
            
            self.status(f"Carregados {len(clientes)} clientes")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao carregar clientes: {e}")
    
    def salvar_cliente(self):
        nome = self.input_nome.text().strip()
        cnpj_cpf = self.input_cnpj_cpf.text().strip()
        endereco = self.input_endereco.text().strip()
        telefone = self.input_telefone.text().strip()
        email = self.input_email.text().strip()
        
        if not nome:
            QMessageBox.warning(self, "Aviso", "O nome é obrigatório!")
            self.input_nome.setFocus()
            return
        
        if email and not self.validar_email(email):
            QMessageBox.warning(self, "Aviso", "Email inválido!")
            self.input_email.setFocus()
            return
        
        try:
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            
            if self.cliente_selecionado:
                cursor.execute('''
                    UPDATE clientes 
                    SET nome=?, cnpj_cpf=?, endereco=?, telefone=?, email=?
                    WHERE id=?
                ''', (nome, cnpj_cpf, endereco, telefone, email, self.cliente_selecionado))
                mensagem = "Cliente atualizado com sucesso!"
            else:
                cursor.execute('''
                    INSERT INTO clientes (nome, cnpj_cpf, endereco, telefone, email)
                    VALUES (?, ?, ?, ?, ?)
                ''', (nome, cnpj_cpf, endereco, telefone, email))
                mensagem = "Cliente cadastrado com sucesso!"
            
            conn.commit()
            conn.close()
            
            QMessageBox.information(self, "Sucesso", mensagem)
            self.limpar_formulario()
            self.carregar_clientes()
            
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Aviso", "CNPJ/CPF já cadastrado!")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao salvar cliente: {e}")
    
    def excluir_cliente(self):
        if not self.cliente_selecionado:
            QMessageBox.warning(self, "Aviso", "Selecione um cliente para excluir!")
            return
        
        resposta = QMessageBox.question(
            self, "Confirmar Exclusão",
            "Tem certeza que deseja excluir este cliente?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if resposta == QMessageBox.StandardButton.Yes:
            try:
                conn = sqlite3.connect('usina_concreto.db')
                cursor = conn.cursor()
                cursor.execute("DELETE FROM clientes WHERE id=?", (self.cliente_selecionado,))
                conn.commit()
                conn.close()
                
                QMessageBox.information(self, "Sucesso", "Cliente excluído com sucesso!")
                self.limpar_formulario()
                self.carregar_clientes()
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Erro ao excluir cliente: {e}")
    
    def selecionar_cliente(self):
        linha = self.tabela_clientes.currentRow()
        if linha >= 0:
            self.cliente_selecionado = int(self.tabela_clientes.item(linha, 0).text())
            self.input_nome.setText(self.tabela_clientes.item(linha, 1).text() or "")
            self.input_cnpj_cpf.setText(self.tabela_clientes.item(linha, 2).text() or "")
            self.input_endereco.setText(self.tabela_clientes.item(linha, 3).text() or "")
            self.input_telefone.setText(self.tabela_clientes.item(linha, 4).text() or "")
            self.input_email.setText(self.tabela_clientes.item(linha, 5).text() or "")
    
    def novo_cliente(self):
        self.limpar_formulario()
        self.input_nome.setFocus()
    
    def limpar_formulario(self):
        self.cliente_selecionado = None
        self.input_nome.clear()
        self.input_cnpj_cpf.clear()
        self.input_endereco.clear()
        self.input_telefone.clear()
        self.input_email.clear()
        self.tabela_clientes.clearSelection()
    
    def validar_email(self, email):
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None
    
    # ------------------ EXPORTAÇÃO ------------------ #
    
    def exportar_clientes(self):
        try:
            import csv
            from datetime import datetime
            
            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM clientes ORDER BY nome")
            clientes = cursor.fetchall()
            conn.close()
            
            if not clientes:
                QMessageBox.warning(self, "Aviso", "Não há clientes para exportar!")
                return
            
            filename = f"clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile, delimiter=';')
                writer.writerow(['ID', 'Nome', 'CNPJ/CPF', 'Endereço', 'Telefone', 'Email', 'Data Cadastro'])
                
                for cliente in clientes:
                    writer.writerow(cliente)
            
            QMessageBox.information(self, "Sucesso", f"Clientes exportados para: {filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar clientes: {e}")
    
    def exportar_clientes_excel(self):
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from datetime import datetime

            conn = sqlite3.connect('usina_concreto.db')
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM clientes ORDER BY nome")
            clientes = cursor.fetchall()
            conn.close()

            if not clientes:
                QMessageBox.warning(self, "Aviso", "Não há clientes para exportar!")
                return

            filename = f"clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Clientes"

            headers = ['ID', 'Nome', 'CNPJ/CPF', 'Endereço', 'Telefone', 'Email', 'Data Cadastro']
            sheet.append(headers)

            for cliente in clientes:
                sheet.append(cliente)

            for i, col in enumerate(headers, 1):
                sheet.column_dimensions[get_column_letter(i)].width = 20

            # Mostra caminho completo e abre o arquivo
            print("Salvando arquivo em:", os.path.abspath(filename))
            workbook.save(filename)
            QMessageBox.information(self, "Sucesso", f"Clientes exportados para: {filename}")

            # Abrir Excel automaticamente no Windows
            if os.name == 'nt':
                subprocess.Popen(['start', '', os.path.abspath(filename)], shell=True)

        except ImportError:
            QMessageBox.warning(self, "Erro", "Instale o módulo 'openpyxl' para exportar para Excel: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao exportar clientes: {e}")
    
    # ------------------ IMPRESSÃO ------------------ #
    
    def imprimir_clientes(self):
        try:
            from PyQt6.QtPrintSupport import QPrinter, QPrintDialog
            from PyQt6.QtGui import QTextDocument

            if self.tabela_clientes.rowCount() == 0:
                QMessageBox.warning(self, "Aviso", "Não há clientes para imprimir!")
                return

            html = "<h2>Lista de Clientes</h2><table border='1' cellspacing='0' cellpadding='4'>"
            html += "<tr>"
            for col in range(self.tabela_clientes.columnCount()):
                html += f"<th>{self.tabela_clientes.horizontalHeaderItem(col).text()}</th>"
            html += "</tr>"

            for row in range(self.tabela_clientes.rowCount()):
                html += "<tr>"
                for col in range(self.tabela_clientes.columnCount()):
                    html += f"<td>{self.tabela_clientes.item(row, col).text()}</td>"
                html += "</tr>"
            html += "</table>"

            doc = QTextDocument()
            doc.setHtml(html)

            printer = QPrinter()
            dialog = QPrintDialog(printer, self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                doc.print(printer)

        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Erro ao imprimir clientes: {e}")
    
    # ------------------ STATUS ------------------ #
    
    def status(self, mensagem):
        if self.parent and hasattr(self.parent, 'status_label'):
            self.parent.status_label.setText(mensagem)